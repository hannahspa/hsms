r"""Import MySpa service catalog into HSMS.

Reads D:\Hannah Spa\Database\danh_sach_dich_vu_tat_ca_chi_nhanh_1778309994.xlsx
and upserts services by ma_dv, falling back to exact service name.
"""
import io
import json
import re
import sys
import urllib.parse
import urllib.request
from datetime import datetime
from pathlib import Path

import openpyxl

sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8")

ROOT = Path(r"C:\Users\Quoc Nam\Projects\hsms")
DB_FILE = Path(r"D:\Hannah Spa\Database\danh_sach_dich_vu_tat_ca_chi_nhanh_1778309994.xlsx")
SUPABASE_URL = "https://aqyemkfbjqxpegingoil.supabase.co"

MYSPA_CATEGORY_REMAP = {
    "lASER": "Công Nghệ Cao - Laser",
    "Dịch Vụ Đồng GIá 29k": "Gội Đầu",
    "Đắp Mặt Nạ": "Chăm Sóc Da Mặt",
}

MISSING_CATEGORY_BY_CODE = {
    "DV-00296": "Combo Khuyến Mãi",
    "DV-00289": "Phụ Thu Dịch Vụ",
    "DV-00287": "Massage Body",
    "DV-00275": "Công Nghệ Cao - Laser",
    "DV-00274": "Công Nghệ Cao - Laser",
    "DV-00273": "Phụ Thu Dịch Vụ",
    "DV-00271": "Phụ Thu Dịch Vụ",
    "DV-00270": "Combo Khuyến Mãi",
    "DV-00268": "Combo Khuyến Mãi",
    "DV-00267": "Phụ Thu Dịch Vụ",
}


def load_service_key():
    text = (ROOT / "scripts" / "import_myspa.py").read_text(encoding="utf-8", errors="ignore")
    match = re.search(r'SERVICE_KEY\s*=\s*"([^"]+)"', text)
    if not match:
        raise RuntimeError("Cannot find SERVICE_KEY in scripts/import_myspa.py")
    return match.group(1)


SERVICE_KEY = load_service_key()
HEADERS = {
    "apikey": SERVICE_KEY,
    "Authorization": "Bearer " + SERVICE_KEY,
    "Content-Type": "application/json",
}


def parse_money(value):
    if value is None:
        return 0
    if isinstance(value, (int, float)):
        return int(round(value))
    raw = str(value).strip()
    if not raw or raw.lower() in ("none", "nan"):
        return 0
    raw = raw.replace("đ", "").replace(" ", "").replace(",", "")
    try:
        return int(round(float(raw)))
    except ValueError:
        return 0


def parse_int(value):
    return parse_money(value)


def parse_commission(value, base_price):
    raw = "" if value is None else str(value).strip()
    if not raw or raw.lower() in ("none", "nan"):
        return {"raw": raw or None, "type": None, "amount": 0, "percent": 0}
    if "%" in raw:
        try:
            pct = float(raw.replace("%", "").strip() or 0)
        except ValueError:
            pct = 0
        return {
            "raw": raw,
            "type": "percent" if pct > 0 else None,
            "amount": int(round(base_price * pct / 100)) if base_price else 0,
            "percent": pct,
        }
    amount = parse_money(raw)
    pct = round(amount / base_price * 100, 4) if amount > 0 and base_price > 0 else 0
    return {
        "raw": raw,
        "type": "absolute" if amount > 0 else None,
        "amount": amount,
        "percent": pct,
    }


def request_json(method, path, payload=None):
    data = None if payload is None else json.dumps(payload, ensure_ascii=False).encode("utf-8")
    req = urllib.request.Request(
        SUPABASE_URL + path,
        data=data,
        method=method,
        headers={**HEADERS, "Prefer": "return=representation" if method == "POST" else "return=minimal"},
    )
    with urllib.request.urlopen(req, timeout=60) as res:
        body = res.read().decode("utf-8")
    return json.loads(body) if body else None


def rest_get_all_services():
    rows = []
    start = 0
    step = 1000
    while True:
        req = urllib.request.Request(
            f"{SUPABASE_URL}/rest/v1/dich_vu?select=id,ma_dv,ten,promotion_config&limit={step}&offset={start}",
            headers=HEADERS,
        )
        with urllib.request.urlopen(req, timeout=60) as res:
            batch = json.loads(res.read().decode("utf-8"))
        rows.extend(batch)
        if len(batch) < step:
            break
        start += step
    return rows


def read_source_rows():
    wb = openpyxl.load_workbook(DB_FILE, read_only=True, data_only=True)
    ws = wb.active
    headers = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
    idx = {str(name).strip(): i for i, name in enumerate(headers)}

    def cell(row, name):
        pos = idx.get(name)
        return row[pos] if pos is not None and pos < len(row) else None

    rows = []
    for raw in ws.iter_rows(min_row=2, values_only=True):
        ma_dv = str(cell(raw, "Mã DV") or "").strip()
        ten = str(cell(raw, "Tên dịch vụ") or "").strip()
        if not ma_dv or not ten:
            continue
        base_price = parse_money(cell(raw, "Số tiền dịch vụ"))
        sale_price = parse_money(cell(raw, "Giá ưu đãi"))
        source_category = str(cell(raw, "Danh mục") or "").strip() or None
        category = MYSPA_CATEGORY_REMAP.get(source_category, source_category)
        if not category:
            category = MISSING_CATEGORY_BY_CODE.get(ma_dv)
        status = str(cell(raw, "Trạng thái (0: Active - 1: Block)") or "0").strip()
        ktv = parse_commission(cell(raw, "Kỹ Thuật Viên-COMMISSION-11"), base_price)
        le_tan = parse_commission(cell(raw, "Lễ Tân-COMMISSION-3"), base_price)
        rows.append({
            "ma_dv": ma_dv,
            "ma_dv_2": str(cell(raw, "Mã DV 2") or "").strip() or None,
            "ten": ten,
            "thoi_gian_phut": parse_int(cell(raw, "Thời lượng")),
            "danh_muc": category,
            "danh_muc_myspa_goc": source_category,
            "gia_co_ban": base_price,
            "gia_uu_dai": sale_price,
            "thu_tu": parse_int(cell(raw, "Ưu tiên")),
            "is_active": status == "0",
            "mo_ta": str(cell(raw, "Mô tả") or "").strip(),
            "chi_nhanh": str(cell(raw, "Chi nhánh") or "").strip() or None,
            "commission_le_tan": le_tan,
            "commission_ktv": ktv,
        })
    wb.close()
    return rows


def build_payload(row, existing=None):
    previous_config = (existing or {}).get("promotion_config") or {}
    if not isinstance(previous_config, dict):
        previous_config = {}
    config = {
        **previous_config,
        "myspa": {
            "source_file": DB_FILE.name,
            "synced_at": datetime.now().isoformat(timespec="seconds"),
            "ma_dv": row["ma_dv"],
            "ma_dv_2": row["ma_dv_2"],
            "chi_nhanh": row["chi_nhanh"],
            "gia_uu_dai": row["gia_uu_dai"],
            "danh_muc_goc": row["danh_muc_myspa_goc"],
            "gia_niem_yet": row["gia_co_ban"],
            "commission_le_tan": row["commission_le_tan"],
            "commission_ktv": row["commission_ktv"],
        },
    }
    ktv_pct = row["commission_ktv"]["percent"] or 0
    return {
        "ma_dv": row["ma_dv"],
        "ten": row["ten"],
        "mo_ta": row["mo_ta"] or "",
        "gia_co_ban": row["gia_co_ban"],
        "ti_le_hoa_hong": ktv_pct,
        "is_active": row["is_active"],
        "danh_muc": row["danh_muc"],
        "nhom_hien_thi": row["danh_muc"],
        "la_phu_thu": row["danh_muc"] == "Phụ Thu Dịch Vụ" or row["ten"].lower().startswith("phụ thu"),
        "thoi_gian_phut": row["thoi_gian_phut"],
        "thu_tu": row["thu_tu"],
        "hien_tren_menu": row["is_active"],
        "promotion_config": config,
    }


def main(dry_run=False):
    source = read_source_rows()
    existing = rest_get_all_services()
    by_code = {r.get("ma_dv"): r for r in existing if r.get("ma_dv")}
    by_name = {r.get("ten"): r for r in existing if r.get("ten")}
    stats = {"source": len(source), "updated": 0, "inserted": 0, "unchanged": 0}
    category_counts = {}
    missing_category = []

    for row in source:
        category_counts[row["danh_muc"] or "-"] = category_counts.get(row["danh_muc"] or "-", 0) + 1
        if not row["danh_muc"]:
            missing_category.append((row["ma_dv"], row["ten"]))
        current = by_code.get(row["ma_dv"]) or by_name.get(row["ten"])
        payload = build_payload(row, current)
        if dry_run:
            stats["updated" if current else "inserted"] += 1
            continue
        if current:
            filter_id = urllib.parse.quote(str(current["id"]), safe="")
            request_json("PATCH", f"/rest/v1/dich_vu?id=eq.{filter_id}", payload)
            stats["updated"] += 1
        else:
            request_json("POST", "/rest/v1/dich_vu", payload)
            stats["inserted"] += 1

    print(json.dumps({
        "file": str(DB_FILE),
        "dry_run": dry_run,
        "stats": stats,
        "categories": category_counts,
        "missing_category_count": len(missing_category),
        "missing_category_sample": missing_category[:20],
    }, ensure_ascii=False, indent=2))


if __name__ == "__main__":
    main("--dry-run" in sys.argv)
