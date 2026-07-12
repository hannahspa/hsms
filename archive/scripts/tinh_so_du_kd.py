import os
"""Tinh so du TK Khanh Duy + phan tich luan chuyen tien quet the"""
import sys
sys.stdout.reconfigure(encoding='utf-8')
import openpyxl
from supabase import create_client
from collections import defaultdict
from datetime import datetime

SUPABASE_URL = 'https://aqyemkfbjqxpegingoil.supabase.co'
SUPABASE_KEY = os.environ["SUPABASE_KEY"]
supabase = create_client(SUPABASE_URL, SUPABASE_KEY)

def fmt(n):
    if n is None: return '0'
    return f'{int(n):,}'.replace(',', '')

def parse_mb(filepath):
    wb = openpyxl.load_workbook(filepath, data_only=True)
    ws = wb['Lịch sử giao dịch']
    txns = []
    for r in range(20, ws.max_row + 1):
        date_val = ws.cell(row=r, column=5).value
        debit = ws.cell(row=r, column=10).value
        credit = ws.cell(row=r, column=11).value
        desc = ws.cell(row=r, column=12).value
        if date_val is None: continue
        try:
            dt = datetime.strptime(str(date_val).strip().split(' ')[0], '%d/%m/%Y')
            date_iso = dt.strftime('%Y-%m-%d')
        except: continue
        no_val = int(float(str(debit).replace(',',''))) if debit else 0
        co_val = int(float(str(credit).replace(',',''))) if credit else 0
        nd = str(desc).strip() if desc else ''
        txns.append({'date': date_iso, 'no': no_val, 'co': co_val, 'desc': nd})
    return txns

base = r'D:\Hannah Spa\Thu Chi\Kiem Toan'
kd = parse_mb(f'{base}\\KhanhDuy.xlsx')
qn = parse_mb(f'{base}\\QuocNam.xlsx')

# ============================================================
print('=' * 75)
print('SO DU TK KHANH DUY (0379080909) — DEN 30/04/2026')
print('=' * 75)

total_no_kd = sum(t['no'] for t in kd)
total_co_kd = sum(t['co'] for t in kd)

# Phan loai tien vao
co_by_type = defaultdict(int)
for t in kd:
    nd = t['desc'].lower()
    if t['co'] == 0: continue
    if 'tip em' in nd: co_by_type['Tip em nhan'] += t['co']
    elif 'hoan lai' in nd or 'hoan tra' in nd or 'tra lai' in nd: co_by_type['Hoan tien/tra lai'] += t['co']
    elif 'dt tm' in nd or 'tm dt' in nd or 'nop tien mat' in nd: co_by_type['Nop tien mat vao TK'] += t['co']
    elif 'quy' in nd or 'quỹ' in nd: co_by_type['Chuyen quy noi bo'] += t['co']
    elif 'lai' in nd and 'gui' in nd: co_by_type['Lai ngan hang'] += t['co']
    elif 'quoc nam' in nd or 'cao quoc' in nd: co_by_type['Tu TK Quoc Nam'] += t['co']
    else: co_by_type['Khach CK (doanh thu)'] += t['co']

# Phan loai tien ra
no_by_type = defaultdict(int)
for t in kd:
    nd = t['desc'].lower()
    if t['no'] == 0: continue
    if 'tip em' in nd: no_by_type['Chuyen tip em'] += t['no']
    elif 'luong' in nd or 'lương' in nd: no_by_type['Luong'] += t['no']
    elif 'ky quy' in nd or 'ký quỹ' in nd: no_by_type['Ky quy'] += t['no']
    elif 'ung luong' in nd or 'ứng lương' in nd: no_by_type['Ung luong'] += t['no']
    elif 'quy' in nd or 'quỹ' in nd: no_by_type['Chuyen quy di'] += t['no']
    elif 'dong hui' in nd or 'đóng hụi' in nd or 'hui' in nd: no_by_type['Dong hui'] += t['no']
    elif 'thanh toan tien hang' in nd: no_by_type['Thanh toan tien hang'] += t['no']
    elif 'mua' in nd or 'may' in nd or 'máy' in nd: no_by_type['Mua sam'] += t['no']
    else: no_by_type['Chi khac (mat bang, dien,...)'] += t['no']

print(f'\n--- TIEN VAO TK KHANH DUY ---')
for k, v in sorted(co_by_type.items(), key=lambda x: -x[1]):
    print(f'  + {k:<35s} {fmt(v)}')

print(f'\n--- TIEN RA TK KHANH DUY ---')
for k, v in sorted(no_by_type.items(), key=lambda x: -x[1]):
    print(f'  - {k:<35s} {fmt(v)}')

so_du_kd = total_co_kd - total_no_kd
print(f'\n  {"="*50}')
print(f'  SO DU TK KHANH DUY (theo sao ke): {fmt(so_du_kd)}')
print(f'  {"="*50}')

# ============================================================
print(f'\n{"="*75}')
print('DOI CHIEU SO DU — SAO KE vs HSMS')
print(f'{"="*75}')

r = supabase.from_('so_du_vi_thuc_te').select('ten,so_du_hien_tai').order('thu_tu').execute()
for vi in (r.data or []):
    print(f'  {vi["ten"]:12s} {fmt(vi.get("so_du_hien_tai",0) or 0)}')

hsms_mb = next((v.get('so_du_hien_tai',0) or 0 for v in (r.data or []) if v['ten'] == 'MB Bank'), 0)
print(f'\n  Chenh lech: Sao ke TK KD ({fmt(so_du_kd)}) vs HSMS MB Bank ({fmt(hsms_mb)})')
print(f'  = {fmt(so_du_kd - hsms_mb)}')

# ============================================================
# Giai thich chenh lech
# ============================================================
print(f'\n{"="*75}')
print('GIAI THICH CHENH LECH')
print(f'{"="*75}')

# HSMS MB Bank = DT_CK - CP_CK + CK_noi_bo vao - CK_noi_bo ra
# Nhung TK Khanh Duy bao gom CA doanh thu CK + nop tien mat + chuyen quy + hoan tien + tip...
# Nen so du TK se > HSMS MB Bank

# Tinh toan cac khoan khong phai doanh thu trong TK KD
non_revenue_in = (co_by_type.get('Nop tien mat vao TK', 0) +
                  co_by_type.get('Chuyen quy noi bo', 0) +
                  co_by_type.get('Hoan tien/tra lai', 0) +
                  co_by_type.get('Tip em nhan', 0) +
                  co_by_type.get('Lai ngan hang', 0) +
                  co_by_type.get('Tu TK Quoc Nam', 0))

non_expense_out = (no_by_type.get('Chuyen tip em', 0) +
                   no_by_type.get('Chuyen quy di', 0) +
                   no_by_type.get('Ung luong', 0))

print(f'\n  Tien vao khong phai doanh thu: {fmt(non_revenue_in)}')
for k, v in sorted(co_by_type.items(), key=lambda x: -x[1]):
    if k != 'Khach CK (doanh thu)':
        print(f'    + {k}: {fmt(v)}')

print(f'\n  Tien ra khong phai chi phi: {fmt(non_expense_out)}')
print(f'\n  Doanh thu CK thuc te trong TK: ~{fmt(total_co_kd - non_revenue_in)}')

# ============================================================
# PHAN TICH QUET THE
# ============================================================
print(f'\n{"="*75}')
print('PHAN TICH TIEN QUET THE')
print(f'{"="*75}')

r = supabase.from_('doanh_thu').select('so_tien').eq('hinh_thuc','quet_the').execute()
hsms_qt = sum(d['so_tien'] or 0 for d in (r.data or []))
print(f'\n  Doanh thu Quet The (HSMS): {fmt(hsms_qt)}')
print(f'  Doanh thu Quet The (MySpa): 77.684.000d')

# Tien quet the duoc TP Bank giu va chuyen ve MB Bank sau 3-7 ngay
# Trong sao ke MB, tien quet the ve se hien thi nhu 1 GD CK tu TP Bank
# Can xac dinh nhung GD nay de biet bao nhieu tien quet the da ve, bao nhieu chua ve

print(f'\n  LUU Y: Tien quet the tu TP Bank chuyen ve MB Bank se hien thi')
print(f'  nhu GD "Khach CK" thong thuong trong sao ke MB. Hien tai chua')
print(f'  tach duoc GD nao la quet the ve, GD nao la khach CK that.')
print(f'')
print(f'  De biet chinh xac: Can sao ke TP Bank de doi chieu.')
print(f'  Tong quet the 77.684.000d - neu da ve het thi da nam trong')
print(f'  so du TK MB la {fmt(so_du_kd)}.')
