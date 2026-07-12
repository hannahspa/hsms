import io
import json
import re
import requests
import subprocess
import sys
from collections import defaultdict
from datetime import datetime

import openpyxl

sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8")

WORKBOOK = r"D:\Hannah Spa\Database\danh_sach_ban_hang_tat_ca_chi_nhanh_1778311594.xlsx"
CUTOFF = "2026-04-30"
BATCH_SIZE = 1000
SUPABASE_URL = "https://aqyemkfbjqxpegingoil.supabase.co"


def sql_quote(value):
    if value is None:
        return "NULL"
    return "'" + str(value).replace("'", "''") + "'"


def load_service_key():
    script_path = r"C:\Users\Quoc Nam\Projects\hsms\scripts\import_orders.py"
    text = open(script_path, "r", encoding="utf-8", errors="ignore").read()
    match = re.search(r'SERVICE_KEY\s*=\s*"([^"]+)"', text)
    if not match:
      raise RuntimeError("Missing Supabase service key. Set SUPABASE_SERVICE_ROLE_KEY or keep scripts/import_orders.py available.")
    return match.group(1)


def parse_money(value):
    if value is None:
        return 0
    if isinstance(value, (int, float)):
        return int(value)
    raw = str(value).strip().replace(",", "").replace(".00", "")
    raw = re.sub(r"[^\d\-]", "", raw)
    try:
        return int(raw or 0)
    except ValueError:
        return 0


def parse_date(value):
    if not value:
        return None
    if isinstance(value, datetime):
        return value.date().isoformat()
    raw = str(value).strip()
    for fmt in ("%d/%m/%Y %H:%M:%S", "%d/%m/%Y %H:%M", "%d/%m/%Y"):
        try:
            return datetime.strptime(raw, fmt).date().isoformat()
        except ValueError:
            pass
    return raw[:10] if re.match(r"^\d{4}-\d{2}-\d{2}", raw) else None


def parse_staff(raw):
    if not raw:
        return []
    rows = []
    for part in str(raw).split("|"):
        part = part.strip()
        if ":" not in part:
            continue
        name, amount = part.split(":", 1)
        name = " ".join(name.split())
        if name:
            rows.append({"name": name, "amount": parse_money(amount)})
    return rows


def norm_name(name):
    text = str(name or "").lower().strip()
    text = re.sub(r"\s+", " ", text)
    return text


def short_staff_name(name):
    text = re.sub(r"\(\s*Nghỉ Việc\s*\)", "", str(name or ""), flags=re.IGNORECASE).strip()
    parts = text.split()
    return " ".join(parts[-2:]) if len(parts) > 2 else " ".join(parts)


def run_sql(sql):
    cmd = ["supabase", "db", "query", "--linked", sql]
    result = subprocess.run(
        cmd,
        text=True,
        encoding="utf-8",
        errors="replace",
        capture_output=True,
        cwd=r"C:\Users\Quoc Nam\Projects\hsms",
    )
    if result.returncode != 0:
        print(result.stdout)
        print(result.stderr)
        raise SystemExit(result.returncode)
    return result.stdout


def load_staff_map():
    out = run_sql("select id, ho_ten, trang_thai from nhan_vien;")
    payload, _ = json.JSONDecoder().raw_decode(out[out.index("{"):])
    rows = payload.get("rows", [])
    exact = {}
    last2 = {}
    for row in rows:
        name = norm_name(row["ho_ten"])
        exact[name] = row["id"]
        parts = name.split()
        if len(parts) >= 2:
            last2[" ".join(parts[-2:])] = row["id"]
    return exact, last2


def match_staff(name, exact, last2):
    key = norm_name(name)
    if key in exact:
        return exact[key]
    parts = key.split()
    if len(parts) >= 2:
        return last2.get(" ".join(parts[-2:]))
    return None


def load_workbook_rows(exact, last2):
    wb = openpyxl.load_workbook(WORKBOOK, read_only=True, data_only=True)
    ws = wb.active
    rows_iter = ws.iter_rows(values_only=True)
    header_values = next(rows_iter)
    headers = {str(value).strip(): idx for idx, value in enumerate(header_values) if value is not None}

    def val(row, col):
        idx = headers.get(col)
        return row[idx] if idx is not None and idx < len(row) else None

    per_order_line = defaultdict(int)
    rows = []
    current_order = None
    current_date = None
    current_header_staff = []

    for idx, row in enumerate(rows_iter, start=2):
        ma_don = str(val(row, "Mã đơn hàng") or "").strip()
        if ma_don and ma_don.lower() not in ("none", "nan"):
            current_order = ma_don
            if parse_date(val(row, "Ngày giờ")):
                current_date = parse_date(val(row, "Ngày giờ"))
            if parse_money(val(row, "Thành tiền ĐH/TLT")) > 0 or not str(val(row, "Mã DV/SP") or "").strip():
                current_header_staff = parse_staff(val(row, "Commission nhân viên"))

        if not current_order or not current_date or current_date > CUTOFF:
            continue

        ma_dv = str(val(row, "Mã DV/SP") or "").strip()
        ten_dv = str(val(row, "Tên DV/SP") or "").strip()
        nhom_dv = str(val(row, "Nhóm DV/SP") or "").strip()
        if not ma_dv or ma_dv.lower() in ("none", "nan") or not ten_dv:
            continue

        staff_rows = parse_staff(val(row, "Commission nhân viên"))
        commission_kind = "line"
        if not staff_rows and (nhom_dv == "Thẻ liệu trình" or ma_dv.startswith("THE-LT")):
            staff_rows = current_header_staff
            commission_kind = "order_card_sale"
        per_order_line[current_order] += 1
        line_no = per_order_line[current_order]

        if not staff_rows:
            rows.append({
                "ma_don": current_order,
                "line_no": line_no,
                "ngay": current_date,
                "ten_dich_vu": ten_dv,
                "item_code": ma_dv,
                "item_group": nhom_dv,
                "staff_name": None,
                "staff_display": None,
                "staff_status": "chua_co_ten",
                "matched_nhan_vien_id": None,
                "commission_amount": 0,
                "raw_commission": None,
                "commission_kind": commission_kind,
            })
            continue

        first = staff_rows[0]
        matched_id = match_staff(first["name"], exact, last2)
        rows.append({
            "ma_don": current_order,
            "line_no": line_no,
            "ngay": current_date,
            "ten_dich_vu": ten_dv,
            "item_code": ma_dv,
            "item_group": nhom_dv,
            "staff_name": first["name"],
            "staff_display": first["name"] if matched_id else f'{short_staff_name(first["name"])} (Nghỉ Việc)',
            "staff_status": "dang_lam" if matched_id else "nghi_viec",
            "matched_nhan_vien_id": matched_id,
            "commission_amount": first["amount"],
            "raw_commission": str(val(row, "Commission nhân viên") or (val(row, "Mã DV/SP") and "HEADER: " + "|".join([f"{s['name']}: {s['amount']}" for s in current_header_staff])) or ""),
            "commission_kind": commission_kind,
        })

        if idx % 25000 == 0:
            print(f"  read {idx:,} worksheet rows...")

    wb.close()
    return rows


def insert_batches(rows):
    run_sql("truncate table myspa_legacy_staff_sync;")
    key = load_service_key()
    headers = {
        "apikey": key,
        "Authorization": "Bearer " + key,
        "Content-Type": "application/json",
        "Prefer": "resolution=merge-duplicates,return=minimal",
    }
    total = len(rows)
    for i in range(0, total, BATCH_SIZE):
        batch = rows[i:i + BATCH_SIZE]
        payload = []
        for row in batch:
            payload.append({
                "ma_don": row["ma_don"],
                "line_no": row["line_no"],
                "ngay": row["ngay"],
                "ten_dich_vu": row["ten_dich_vu"],
                "item_code": row["item_code"],
                "item_group": row["item_group"],
                "staff_name": row["staff_name"],
                "staff_display": row["staff_display"],
                "staff_status": row["staff_status"],
                "matched_nhan_vien_id": row["matched_nhan_vien_id"],
                "commission_amount": row["commission_amount"],
                "raw_commission": row["raw_commission"],
                "commission_kind": row["commission_kind"],
                "cutoff_date": CUTOFF,
            })
        response = requests.post(
            SUPABASE_URL + "/rest/v1/myspa_legacy_staff_sync?on_conflict=ma_don,line_no",
            headers=headers,
            json=payload,
            timeout=120,
        )
        if not response.ok:
            raise RuntimeError(f"Stage failed: {response.status_code} {response.text[:500]}")
        print(f"  staged {min(i + BATCH_SIZE, total):,}/{total:,}")


def apply_sync():
    sql = f"""
    with numbered as (
      select
        dhct.id,
        dhct.nhan_vien_id,
        dhct.meta,
        dhct.loai_item,
        dhct.tien_hoa_hong,
        dhct.tien_tour,
        dhct.tien_commission,
        row_number() over (partition by dh.ma_don order by dhct.created_at, dhct.id) as line_no,
        dh.ma_don,
        dh.ngay
      from don_hang_chi_tiet dhct
      join don_hang dh on dh.id = dhct.don_hang_id
      where dh.ngay <= date '{CUTOFF}'
    ),
    matched as (
      select
        n.id,
        s.staff_name,
        s.staff_display,
        s.staff_status,
        s.matched_nhan_vien_id,
        s.commission_amount,
        s.raw_commission,
        s.ten_dich_vu,
        s.item_code,
        s.item_group,
        s.commission_kind
      from numbered n
      join myspa_legacy_staff_sync s
        on s.ma_don = n.ma_don
       and s.line_no = n.line_no
    )
    update don_hang_chi_tiet dhct
    set
      nhan_vien_id = coalesce(dhct.nhan_vien_id, matched.matched_nhan_vien_id),
      loai_item = case
        when matched.commission_kind = 'order_card_sale' and dhct.loai_item = 'dich_vu' then 'the_moi'
        else dhct.loai_item
      end,
      tien_hoa_hong = case
        when matched.commission_kind = 'order_card_sale' and coalesce(dhct.tien_hoa_hong, 0) = 0 then matched.commission_amount
        else dhct.tien_hoa_hong
      end,
      tien_tour = case
        when matched.commission_kind = 'order_card_sale' then 0
        else dhct.tien_tour
      end,
      tien_commission = case
        when matched.commission_kind = 'order_card_sale' then matched.commission_amount
        else dhct.tien_commission
      end,
      meta = coalesce(dhct.meta, '{{}}'::jsonb) || jsonb_build_object(
        'myspaStaffName', matched.staff_name,
        'myspaStaffDisplay', matched.staff_display,
        'myspaStaffStatus', matched.staff_status,
        'myspaStaffCommission', matched.commission_amount,
        'myspaStaffRaw', matched.raw_commission,
        'myspaItemCode', matched.item_code,
        'myspaItemGroup', matched.item_group,
        'myspaCommissionKind', matched.commission_kind,
        'tenDichVu', case when matched.commission_kind = 'order_card_sale' then matched.ten_dich_vu else coalesce(dhct.meta->>'tenDichVu', matched.ten_dich_vu) end,
        'myspaCutoffDate', '{CUTOFF}'
      )
    from matched
    where dhct.id = matched.id;

    insert into nhan_vien_thu_nhap (
      don_hang_id,
      don_hang_chi_tiet_id,
      nhan_vien_id,
      loai,
      nguon,
      ngay,
      doanh_so_tinh,
      ti_le,
      so_tien,
      trang_thai,
      is_test,
      ghi_chu
    )
    select
      dh.id,
      dhct.id,
      dhct.nhan_vien_id,
      case
        when dhct.loai_item in ('dich_vu','the_lieu_trinh') then 'tour'
        when dhct.loai_item in ('san_pham','the_moi') then 'commission'
        else 'tour'
      end,
      'pos',
      dh.ngay,
      coalesce(dhct.thanh_tien, 0),
      dhct.ti_le_hoa_hong,
      case
        when dhct.loai_item in ('dich_vu','the_lieu_trinh') then coalesce(nullif(dhct.tien_tour, 0), dhct.tien_hoa_hong, 0)
        when dhct.loai_item in ('san_pham','the_moi') then coalesce(nullif(dhct.tien_commission, 0), dhct.tien_hoa_hong, 0)
        else coalesce(dhct.tien_hoa_hong, 0)
      end,
      'doi_soat',
      coalesce(dh.is_test, false),
      'Đồng bộ tên nhân viên từ MySpa đến 30/04/2026'
    from don_hang_chi_tiet dhct
    join don_hang dh on dh.id = dhct.don_hang_id
    where dh.ngay <= date '{CUTOFF}'
      and dh.trang_thai <> 'huy'
      and dhct.nhan_vien_id is not null
      and (
        (dhct.loai_item in ('dich_vu','the_lieu_trinh') and coalesce(nullif(dhct.tien_tour, 0), dhct.tien_hoa_hong, 0) > 0)
        or (dhct.loai_item in ('san_pham','the_moi') and coalesce(nullif(dhct.tien_commission, 0), dhct.tien_hoa_hong, 0) > 0)
      )
      and not exists (
        select 1
        from nhan_vien_thu_nhap nvt
        where nvt.don_hang_chi_tiet_id = dhct.id
          and nvt.loai = case
            when dhct.loai_item in ('dich_vu','the_lieu_trinh') then 'tour'
            when dhct.loai_item in ('san_pham','the_moi') then 'commission'
            else 'tour'
          end
      );
    """
    run_sql(sql)


def main():
    print("Loading staff map...")
    exact, last2 = load_staff_map()
    print("Reading MySpa workbook...")
    rows = load_workbook_rows(exact, last2)
    print(f"Rows to stage: {len(rows):,}")
    print("Staging legacy staff rows...")
    insert_batches(rows)
    print("Applying sync to HSMS line items...")
    apply_sync()
    print(run_sql("select staff_status, count(*) as rows, coalesce(sum(commission_amount),0) as amount from myspa_legacy_staff_sync group by staff_status order by rows desc;"))
    print("Done.")


if __name__ == "__main__":
    main()
