r"""Read-only audit for locking MySpa legacy data up to 2026-04-30.

This script compares the exported MySpa files in D:\Hannah Spa\Database with
the HSMS Supabase data. It does not write to Supabase.
"""
import json
import re
import urllib.parse
import urllib.request
from datetime import datetime
from pathlib import Path

import openpyxl

ROOT = Path(r"C:\Users\Quoc Nam\Projects\hsms")
DB = Path(r"D:\Hannah Spa\Database")
CUTOFF = "2026-04-30"
SUPABASE_URL = "https://aqyemkfbjqxpegingoil.supabase.co"

FILES = {
    "customers": DB / "khach_hang_tat_ca_chi_nhanh_1778309550.xlsx",
    "cards": DB / "danh_sach_the_lieu_trinh_tat_ca_chi_nhanh_1778310007.xlsx",
    "sales": DB / "danh_sach_ban_hang_tat_ca_chi_nhanh_1778311594.xlsx",
    "services": DB / "danh_sach_dich_vu_tat_ca_chi_nhanh_1778309994.xlsx",
    "combos": DB / "danh_sach_combo_lieu_trinh_tat_ca_chi_nhanh_1778310109.xlsx",
}


def load_service_key():
    text = (ROOT / "scripts/import_myspa.py").read_text(encoding="utf-8", errors="ignore")
    match = re.search(r'SERVICE_KEY\s*=\s*"([^"]+)"', text)
    if not match:
        raise RuntimeError("Cannot find SERVICE_KEY in scripts/import_myspa.py")
    return match.group(1)


SERVICE_KEY = load_service_key()


def parse_date(value):
    if isinstance(value, datetime):
        return value.date().isoformat()
    if not value:
        return None
    raw = str(value).strip()
    for fmt in ("%d/%m/%Y %H:%M:%S", "%d/%m/%Y %H:%M", "%d/%m/%Y", "%Y-%m-%d %H:%M:%S", "%Y-%m-%d"):
        try:
            text = raw[:19] if "%H" in fmt else raw[:10]
            return datetime.strptime(text, fmt).date().isoformat()
        except ValueError:
            pass
    return None


def read_sheet_rows(path):
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb.active
    headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
    for row in ws.iter_rows(min_row=2, values_only=True):
        if any(value not in (None, "") for value in row):
            yield headers, row
    wb.close()


def rest_get_all(table, columns):
    rows = []
    start = 0
    step = 1000
    encoded_columns = urllib.parse.quote(columns, safe=",()*")
    while True:
        request = urllib.request.Request(
            f"{SUPABASE_URL}/rest/v1/{table}?select={encoded_columns}",
            headers={
                "apikey": SERVICE_KEY,
                "Authorization": "Bearer " + SERVICE_KEY,
                "Range": f"{start}-{start + step - 1}",
            },
        )
        with urllib.request.urlopen(request, timeout=60) as response:
            batch = json.loads(response.read().decode("utf-8"))
        rows.extend(batch)
        if len(batch) < step:
            break
        start += step
    return rows


def source_customers():
    codes = set()
    for _, row in read_sheet_rows(FILES["customers"]):
        code = str(row[0] or "").strip()
        if code:
            codes.add(code)
    return codes


def source_cards():
    cards = {}
    for _, row in read_sheet_rows(FILES["cards"]):
        code = str(row[0] or "").strip()
        created = parse_date(row[1])
        if code and created and created <= CUTOFF:
            cards[code] = {
                "ma_the": code,
                "ngay_tao": created,
                "ma_kh": str(row[2] or "").strip(),
                "ten_kh": str(row[3] or "").strip(),
                "ten_dich_vu": str(row[7] or row[5] or "").strip(),
                "tong_so_lan": row[8],
                "da_su_dung": row[9],
            }
    return cards


def source_orders():
    headers = set()
    with_lines = set()
    line_count = 0
    for _, row in read_sheet_rows(FILES["sales"]):
        order_code = str(row[0] or "").strip()
        date = parse_date(row[1])
        if not order_code or not date or date > CUTOFF:
            continue
        headers.add(order_code)
        item_code = str(row[8] or "").strip()
        item_name = str(row[9] or "").strip()
        if item_code and item_code.lower() not in ("none", "nan") and item_name:
            with_lines.add(order_code)
            line_count += 1
    return headers, with_lines, line_count


def main():
    src_customers = source_customers()
    src_cards = source_cards()
    src_order_headers, src_orders_with_lines, src_order_lines = source_orders()

    db_customers = {row.get("ma_kh") for row in rest_get_all("khach_hang", "ma_kh") if row.get("ma_kh")}
    db_cards = {
        row.get("ma_the")
        for row in rest_get_all("the_lieu_trinh", "ma_the,ngay_mua")
        if row.get("ma_the") and (row.get("ngay_mua") or "9999-99-99") <= CUTOFF
    }
    db_orders = {
        row.get("ma_don")
        for row in rest_get_all("don_hang", "ma_don,ngay,is_test")
        if row.get("ma_don") and (row.get("ngay") or "9999-99-99") <= CUTOFF and not row.get("is_test")
    }

    missing_cards = sorted(set(src_cards) - db_cards)
    report = {
        "cutoff": CUTOFF,
        "exports_available": {name: path.exists() for name, path in FILES.items()},
        "customers": {
            "source": len(src_customers),
            "hsms": len(db_customers),
            "missing_in_hsms": len(src_customers - db_customers),
            "missing_sample": sorted(src_customers - db_customers)[:20],
        },
        "orders": {
            "source_headers": len(src_order_headers),
            "source_orders_with_lines": len(src_orders_with_lines),
            "source_lines": src_order_lines,
            "hsms_orders": len(db_orders),
            "missing_in_hsms": len(src_order_headers - db_orders),
            "extra_in_hsms": len(db_orders - src_order_headers),
            "missing_sample": sorted(src_order_headers - db_orders)[:20],
            "extra_sample": sorted(db_orders - src_order_headers)[:20],
        },
        "treatment_cards": {
            "source": len(src_cards),
            "hsms": len(db_cards),
            "missing_in_hsms": len(missing_cards),
            "missing_sample": [src_cards[code] for code in missing_cards[:20]],
        },
        "needed_realtime_exports": [
            "Báo cáo Commission theo mốc ngày để khóa tiền tour/hoa hồng nhân sự",
            "Danh sách thẻ liệu trình realtime sau 30/04 nếu muốn cập nhật đến hôm nay",
            "Danh sách bán hàng realtime sau 30/04 nếu muốn nối tiếp vận hành",
            "Thẻ trả trước / ví khách hàng nếu vẫn cần đối soát số dư cũ",
            "Lịch hẹn nếu muốn CRM đầy đủ phần đặt hẹn",
        ],
    }
    print(json.dumps(report, ensure_ascii=False, indent=2))


if __name__ == "__main__":
    main()
