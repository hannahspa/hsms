# -*- coding: utf-8 -*-
"""
ĐỐI CHIẾU SỐ LƯỢNG THẺ MỖI KHÁCH: MySpa ↔ HSMS (09/07/2026).

NV lo: sau khi cắt MySpa, nếu HSMS "nhảy thêm thẻ dịch vụ" thì không có gì đối
soát. Script này đối chiếu theo dạng ĐƠN GIẢN NHẤT NV đề xuất: mỗi SĐT khách
sở hữu BAO NHIÊU THẺ ở mỗi hệ → xem lệch chỗ nào (khách nào HSMS dư/thiếu thẻ).

Khác với doi_chieu_the_sdt_dv.py (so theo dịch vụ + buổi) — đây so THUẦN SỐ THẺ
/khách để bắt lỗi "nhảy thêm thẻ".

Chạy:  python scripts/doi_chieu_so_the_moi_khach.py [file_excel]
Xuất:  D:\Hannah Spa\Database\DOI_CHIEU_SO_THE_KHACH_<ngày>.xlsx
       - LECH: khách có số thẻ MySpa ≠ HSMS (cần xem)
       - KHOP: khách khớp số thẻ (chỉ để tham khảo)
CHỈ ĐỌC — không sửa DB.
"""
import sys, io, os, glob, re
from datetime import datetime
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')

import openpyxl
from openpyxl import Workbook
from supabase import create_client
from dotenv import load_dotenv

load_dotenv('.env.import'); load_dotenv('.env')
URL = os.environ.get('SUPABASE_URL') or os.environ.get('VITE_SUPABASE_URL') or 'https://api.hannahspa.vn'
sb = create_client(URL, os.environ['SUPABASE_KEY'])
DB_DIR = r"D:\Hannah Spa\Database"

def chuan_sdt(s):
    d = re.sub(r'\D', '', str(s or ''))
    if d.startswith('84') and len(d) > 9: d = '0' + d[2:]
    if d and not d.startswith('0'): d = '0' + d
    return d if len(d) >= 9 else ''

# ── 1. MySpa: đếm thẻ theo SĐT (chỉ thẻ CÒN GIÁ TRỊ: còn buổi HOẶC còn hạn) ──
excel = next((a for a in sys.argv[1:] if not a.startswith('--')), None) or \
    sorted(glob.glob(os.path.join(DB_DIR, 'danh_sach_the_lieu_trinh_tat_ca_chi_nhanh_*.xlsx')))[-1]
print(f'📂 {excel}')
wb = openpyxl.load_workbook(excel, read_only=True, data_only=True); ws = wb.active

def to_int(x):
    try: return int(float(str(x).replace(',', '')))
    except: return 0

my = {}       # sdt -> list thẻ (ma, dv, con)
my_name = {}
for i, row in enumerate(ws.iter_rows(values_only=True)):
    if i == 0 or not row[0]: continue
    sdt = chuan_sdt(row[4])
    if not sdt: continue
    con = to_int(row[8]) - to_int(row[9])
    my.setdefault(sdt, []).append({'ma': str(row[0]).strip(), 'dv': row[5] or row[7], 'con': con})
    my_name[sdt] = row[3]
print(f'   MySpa: {sum(len(v) for v in my.values())} thẻ / {len(my)} khách có SĐT')

# ── 2. HSMS: đếm thẻ theo SĐT (bỏ thẻ đã xóa mềm) ──
hs, offset = [], 0
while True:
    r = sb.table('the_lieu_trinh').select(
        'ma_the, ten_dich_vu, so_buoi_tong, so_buoi_da_dung, da_xoa, '
        'khach_hang:khach_hang_id(ho_ten, so_dien_thoai)'
    ).range(offset, offset+999).execute()
    hs.extend(r.data or [])
    if len(r.data or []) < 1000: break
    offset += 1000
hsm = {}
for t in hs:
    if t.get('da_xoa'): continue
    sdt = chuan_sdt((t.get('khach_hang') or {}).get('so_dien_thoai'))
    if not sdt: continue
    con = max(0, (t.get('so_buoi_tong') or 0) - (t.get('so_buoi_da_dung') or 0))
    hsm.setdefault(sdt, []).append({'ma': t.get('ma_the'), 'dv': t.get('ten_dich_vu'), 'con': con})
print(f'   HSMS:  {sum(len(v) for v in hsm.values())} thẻ / {len(hsm)} khách có SĐT')

# ── 3. So số thẻ mỗi khách ──
all_sdt = set(my) | set(hsm)
lech, khop = [], 0
for sdt in all_sdt:
    m = my.get(sdt, []); h = hsm.get(sdt, [])
    if len(m) == len(h):
        khop += 1; continue
    lech.append({
        'sdt': sdt, 'ten': my_name.get(sdt) or (h[0] if h else {}),
        'so_the_myspa': len(m), 'so_the_hsms': len(h),
        'chenh': len(h) - len(m),
        'dv_myspa': ' | '.join(f"{x['dv']}({x['con']})" for x in m),
        'dv_hsms':  ' | '.join(f"{x['dv']}({x['con']})" for x in h),
    })

# ── 4. Xuất ──
lech.sort(key=lambda r: -abs(r['chenh']))
out = os.path.join(DB_DIR, f'DOI_CHIEU_SO_THE_KHACH_{datetime.now():%Y%m%d}.xlsx')
wbo = Workbook(); o = wbo.active; o.title = 'LECH SO THE'
cols = ['SĐT', 'Tên KH', 'Số thẻ MySpa', 'Số thẻ HSMS', 'Chênh (HSMS-MySpa)', 'Thẻ bên MySpa', 'Thẻ bên HSMS']
o.append(cols)
for r in lech:
    ten = r['ten'] if isinstance(r['ten'], str) else (r['ten'].get('ho_ten') if isinstance(r['ten'], dict) else '')
    o.append([r['sdt'], ten, r['so_the_myspa'], r['so_the_hsms'], r['chenh'], r['dv_myspa'], r['dv_hsms']])
wbo.save(out)

print(f'''
════════ SỐ THẺ MỖI KHÁCH ════════
  Khách khớp số thẻ : {khop}
  Khách LỆCH số thẻ : {len(lech)}
    · HSMS DƯ thẻ (chênh > 0): {sum(1 for r in lech if r['chenh'] > 0)}
    · HSMS THIẾU (chênh < 0) : {sum(1 for r in lech if r['chenh'] < 0)}
  📄 {out}
''')
for r in lech[:15]:
    ten = r['ten'] if isinstance(r['ten'], str) else ''
    print(f"  {r['sdt']} {ten[:18]:18} MySpa {r['so_the_myspa']} vs HSMS {r['so_the_hsms']}  ({r['chenh']:+d})")
