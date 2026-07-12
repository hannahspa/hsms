import os
"""Kiem toan tong hop tu file MB Bank goc — Khanh Duy + Quoc Nam"""
import sys
sys.stdout.reconfigure(encoding='utf-8')
import openpyxl
from supabase import create_client
from collections import defaultdict
from datetime import datetime
import re

SUPABASE_URL = 'https://aqyemkfbjqxpegingoil.supabase.co'
SUPABASE_KEY = os.environ["SUPABASE_KEY"]
supabase = create_client(SUPABASE_URL, SUPABASE_KEY)

def fmt(n):
    if n is None: return '0d'
    return f'{int(n):,}d'.replace(',', '.')

def parse_mb_excel(filepath):
    """Parse MB Bank export file"""
    wb = openpyxl.load_workbook(filepath, data_only=True)
    ws = wb['Lịch sử giao dịch']

    # Find account number and period
    acct_no = ''
    period = ''
    for row in ws.iter_rows(min_row=8, max_row=15, values_only=True):
        for cell in row:
            if cell and 'Số tài khoản' in str(cell):
                acct_no = str(cell).split(':')[-1].strip()
            if cell and 'Kỳ sao kê' in str(cell):
                period = str(cell).split(':')[-1].strip()

    # Parse transactions - correct MB Bank column positions
    # Col 2=STT, Col 5=Ngay, Col 7=Ref, Col 10=PS No, Col 11=PS Co, Col 12=ND, Col 16=DV, Col 18=TK, Col 22=NH
    txns = []
    for r in range(20, ws.max_row + 1):
        stt = ws.cell(row=r, column=2).value
        date_val = ws.cell(row=r, column=5).value
        ref_no = ws.cell(row=r, column=7).value
        debit = ws.cell(row=r, column=10).value
        credit = ws.cell(row=r, column=11).value
        desc = ws.cell(row=r, column=12).value
        party = ws.cell(row=r, column=16).value
        acct = ws.cell(row=r, column=18).value
        bank = ws.cell(row=r, column=22).value

        if date_val is None:
            continue
        if stt and isinstance(stt, str) and ('STT' in stt or 'No' in stt):
            continue
        if debit is None and credit is None:
            continue

        # Parse date
        date_str = str(date_val).strip()
        try:
            dt = datetime.strptime(date_str.split(' ')[0], '%d/%m/%Y')
            date_iso = dt.strftime('%Y-%m-%d')
        except:
            continue

        # Parse amounts
        def parse_amt(v):
            if v is None: return 0
            if isinstance(v, (int, float)): return int(v)
            return int(float(str(v).replace(',', '')))

        no_val = parse_amt(debit)
        co_val = parse_amt(credit)
        nd = str(desc).strip() if desc else ''
        party_str = str(party).strip() if party else ''
        bank_str = str(bank).strip() if bank else ''
        acct_str = str(acct).strip() if acct else ''

        txns.append({
            'date': date_iso, 'ref': str(ref_no).strip() if ref_no else '',
            'no': no_val, 'co': co_val,
            'nd': nd, 'party': party_str, 'acct': acct_str, 'bank': bank_str,
        })

    return {'acct_no': acct_no, 'period': period, 'txns': txns}

# ============================================================
# DOC DU LIEU
# ============================================================
base = r'D:\Hannah Spa\Thu Chi\Kiem Toan'
kd_data = parse_mb_excel(f'{base}\\KhanhDuy.xlsx')
qn_data = parse_mb_excel(f'{base}\\QuocNam.xlsx')

print('=' * 80)
print('DU LIEU GOC MB BANK')
print('=' * 80)
print(f'\nKHANH DUY: TK {kd_data["acct_no"]} | {kd_data["period"]} | {len(kd_data["txns"])} GD')
print(f'QUOC NAM:  TK {qn_data["acct_no"]} | {qn_data["period"]} | {len(qn_data["txns"])} GD')

# ============================================================
# PHAN LOAI GIAO DICH
# ============================================================
def classify(tx, is_kd=True):
    """Classify transaction as THU (customer payment) or CHI (expense/transfer out)"""
    nd = tx['nd'].lower()
    co = tx['co']
    no = tx['no']

    if co > 0:  # Money coming IN
        if 'hoan lai' in nd or 'hoan tra' in nd or 'tra lai' in nd:
            return 'HOAN_TIEN'
        if 'tip em' in nd:
            return 'TIP_NHAN'
        if is_kd and ('dt tm' in nd or 'tm dt' in nd or 'nop tien mat' in nd):
            return 'NOP_TIEN_MAT'
        if 'lai' in nd and 'gui' in nd:
            return 'LAI_NGAN_HANG'
        # Check if it's from the spa owner (internal)
        if 'quoc nam' in nd.lower() or 'cao quoc' in nd.lower():
            return 'NOI_BO_VAO'
        return 'KHACH_CK'
    else:  # Money going OUT
        if 'tip em' in nd:
            return 'CHUYEN_TIP'
        if 'luong' in nd or 'lương' in nd:
            return 'LUONG'
        if 'ky quy' in nd or 'ký quỹ' in nd or 'ky quy' in nd:
            return 'KY_QUY'
        if 'ung luong' in nd or 'ứng lương' in nd:
            return 'UNG_LUONG'
        if 'quy' in nd or 'quỹ' in nd:
            return 'QUY_CHUYEN'
        if 'hannah' in nd and ('spa' in nd or 'quy' in nd):
            return 'NOI_BO_RA'
        if 'thu no' in nd or 'the tin dung' in nd:
            return 'THU_NO'
        if 'dong hui' in nd or 'đóng hụi' in nd or 'hui' in nd:
            return 'DONG_HUI'
        if 'thanh toan tien hang' in nd:
            return 'THANH_TOAN_HANG'
        if 'mua' in nd or 'máy' in nd or 'hang' in nd:
            return 'MUA_SAM'
        return 'CHI_KHAC'

for tx in kd_data['txns']:
    tx['cat'] = classify(tx, True)
for tx in qn_data['txns']:
    tx['cat'] = classify(tx, False)

# ============================================================
# TONG QUAN TUNG TK
# ============================================================
for label, data in [('KHANH DUY (0379080909)', kd_data), ('QUOC NAM (320011919)', qn_data)]:
    txns = data['txns']
    total_no = sum(t['no'] for t in txns)
    total_co = sum(t['co'] for t in txns)

    print(f'\n{"="*80}')
    print(f'{label} | {len(txns)} GD | {data["period"]}')
    print(f'{"="*80}')
    print(f'  Tong PS No (Chi):  {fmt(total_no)}')
    print(f'  Tong PS Co (Thu):  {fmt(total_co)}')
    print(f'  Chenh lech:        {fmt(total_co - total_no)}')

    # By category
    by_cat = defaultdict(lambda: {'no': 0, 'co': 0, 'count': 0})
    for t in txns:
        by_cat[t['cat']]['no'] += t['no']
        by_cat[t['cat']]['co'] += t['co']
        by_cat[t['cat']]['count'] += 1

    print(f'\n  {"Loai":<20s} {"PS No":>14s} {"PS Co":>14s} {"Count":>6s}')
    for cat, v in sorted(by_cat.items(), key=lambda x: -(x[1]['no'] + x[1]['co'])):
        print(f'  {cat:<20s} {fmt(v["no"]):>14s} {fmt(v["co"]):>14s} {v["count"]:>6d}')

# ============================================================
# DOI CHIEU DOANH THU KHACH CK vs HSMS
# ============================================================
print(f'\n{"="*80}')
print('DOI CHIEU: KHACH CK (SAO KE) vs HSMS DOANH THU CHUYEN KHOAN')
print(f'{"="*80}')

# Collect all KHACH_CK from both accounts
all_khach_ck = []
for t in kd_data['txns']:
    if t['cat'] == 'KHACH_CK':
        all_khach_ck.append(t)
for t in qn_data['txns']:
    if t['cat'] == 'KHACH_CK':
        all_khach_ck.append(t)

sk_by_day = defaultdict(int)
for t in all_khach_ck:
    sk_by_day[t['date']] += t['co']

# Get HSMS data
r = supabase.from_('doanh_thu').select('ngay,so_tien,dien_giai') \
    .eq('hinh_thuc','chuyen_khoan') \
    .order('ngay').execute()

hsms_by_day = defaultdict(int)
hsms_detail = defaultdict(list)
for d in (r.data or []):
    hsms_by_day[d['ngay']] += (d['so_tien'] or 0)
    hsms_detail[d['ngay']].append(d)

all_days = sorted(set(list(sk_by_day.keys()) + list(hsms_by_day.keys())))

# Group by month
print(f'\n  {"Thang":<10s} {"Sao ke":>16s} {"HSMS":>16s} {"Chenh":>14s} {"GD SK":>6s} {"GD HSMS":>6s}')
for month in sorted(set(d[:7] for d in all_days)):
    sk_m = sum(t['co'] for t in all_khach_ck if t['date'][:7] == month)
    hs_m = sum(v for d, v in hsms_by_day.items() if d[:7] == month)
    n_sk = len([t for t in all_khach_ck if t['date'][:7] == month])
    n_hs = sum(1 for d in all_days if d[:7] == month and hsms_by_day[d] > 0)
    diff = sk_m - hs_m
    icon = 'OK' if diff == 0 else f'LECH {fmt(diff)}'
    print(f'  {month:<10s} {fmt(sk_m):>16s} {fmt(hs_m):>16s} {icon:>14s} {n_sk:>6d} {n_hs:>6d}')

total_sk = sum(t['co'] for t in all_khach_ck)
total_hs = sum(v for v in hsms_by_day.values())
print(f'  {"-"*10} {"-"*16} {"-"*16}')
print(f'  {"TONG":<10s} {fmt(total_sk):>16s} {fmt(total_hs):>16s} {fmt(total_sk - total_hs):>14s}')

# ============================================================
# DOI CHIEU THEO TUNG NGAY (CHI TIET)
# ============================================================
print(f'\n{"="*80}')
print('CHI TIET TUNG NGAY CO CHENH LECH > 100.000d')
print(f'{"="*80}')

big_diff_days = []
for day in all_days:
    sk = sk_by_day.get(day, 0)
    hs = hsms_by_day.get(day, 0)
    diff = sk - hs
    if abs(diff) >= 100000:
        big_diff_days.append((day, sk, hs, diff))

big_diff_days.sort(key=lambda x: abs(x[3]), reverse=True)

for day, sk, hs, diff in big_diff_days[:30]:
    print(f'\n--- {day}: Sao ke={fmt(sk)} HSMS={fmt(hs)} Chenh={fmt(diff)} ---')
    sk_txns = [t for t in all_khach_ck if t['date'] == day]
    for t in sk_txns[:5]:
        print(f'  SK: {fmt(t["co"])}  {t["nd"][:80]}')
    if len(sk_txns) > 5:
        print(f'  ... ({len(sk_txns) - 5} more SK txns)')
    hs_txns = hsms_detail.get(day, [])
    for d in hs_txns:
        print(f'  HS: {fmt(d["so_tien"])}  {(d.get("dien_giai") or "")[:80]}')

# ============================================================
# KET LUAN
# ============================================================
print(f'\n{"="*80}')
print('KET LUAN KIEM TOAN')
print(f'{"="*80}')
print(f'''
1. TK KHANH DUY (0379080909): {len(kd_data["txns"])} GD tu 01/01/2026 den 30/04/2026
   - Khach CK: {fmt(sum(t["co"] for t in kd_data["txns"] if t["cat"]=="KHACH_CK"))}

2. TK QUOC NAM (320011919): {len(qn_data["txns"])} GD tu 26/11/2025 den 19/01/2026
   - Khach CK: {fmt(sum(t["co"] for t in qn_data["txns"] if t["cat"]=="KHACH_CK"))}

3. TONG KHACH CK (ca 2 TK): {fmt(total_sk)}
   TONG HSMS DT CK:          {fmt(total_hs)}
   CHENH LECH:               {fmt(total_sk - total_hs)}
''')
