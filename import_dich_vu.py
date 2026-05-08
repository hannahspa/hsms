"""
Import danh sách dịch vụ từ Excel vào Supabase
Chạy: python import_dich_vu.py

Yêu cầu: pip install openpyxl python-dotenv requests
"""

import os
import json
import math
import requests
import openpyxl
from pathlib import Path
from dotenv import load_dotenv

# ── Config ────────────────────────────────────────────────────────────────────
load_dotenv(".env.import")

SUPABASE_URL = os.getenv("SUPABASE_URL", "https://aqyemkfbjqxpegingoil.supabase.co")
SUPABASE_KEY = os.getenv("SUPABASE_KEY")  # service_role key

if not SUPABASE_KEY:
    print("❌ Thiếu SUPABASE_KEY trong .env.import")
    exit(1)

HEADERS = {
    "apikey": SUPABASE_KEY,
    "Authorization": f"Bearer {SUPABASE_KEY}",
    "Content-Type": "application/json",
    "Prefer": "return=representation",
}

# ── Đường dẫn file Excel ──────────────────────────────────────────────────────
# Anh đặt đúng đường dẫn file Excel của mình ở đây:
EXCEL_FILE = Path(r"C:\Users\Quoc Nam\Desktop\danh_sach_commission_tat_ca_chi_nhanh_1777703001.xlsx")

# Nếu file không tìm thấy, thử tìm trong thư mục project:
if not EXCEL_FILE.exists():
    for f in Path(".").glob("*.xlsx"):
        if "dich_vu" in f.name.lower() or "commission" in f.name.lower() or "dich" in f.name.lower():
            EXCEL_FILE = f
            break

print(f"📂 File Excel: {EXCEL_FILE}")
if not EXCEL_FILE.exists():
    print("❌ Không tìm thấy file Excel. Anh sửa đường dẫn EXCEL_FILE ở trên.")
    exit(1)

# ── Đọc Excel ─────────────────────────────────────────────────────────────────
wb = openpyxl.load_workbook(EXCEL_FILE, data_only=True)

# Tìm sheet chứa dịch vụ
sheet = None
for name in wb.sheetnames:
    print(f"   Sheet: {name}")
    ws = wb[name]
    # Tìm sheet có header "Tên dịch vụ"
    for row in ws.iter_rows(min_row=1, max_row=3, values_only=True):
        for cell in row:
            if cell and "tên dịch vụ" in str(cell).lower():
                sheet = ws
                print(f"   ✅ Dùng sheet: {name}")
                break
        if sheet:
            break
    if sheet:
        break

if not sheet:
    # Dùng sheet đầu tiên nếu không tìm được
    sheet = wb.active
    print(f"   ⚠️  Dùng sheet mặc định: {sheet.title}")

# ── Xác định header ───────────────────────────────────────────────────────────
headers_row = None
col_map = {}  # tên_cột → index (1-based)

for row_idx, row in enumerate(sheet.iter_rows(min_row=1, max_row=5, values_only=True), 1):
    row_str = [str(c).lower().strip() if c else "" for c in row]
    # Tìm dòng có "tên dịch vụ"
    for col_idx, cell in enumerate(row_str, 1):
        if "tên dịch vụ" in cell or "ten dich vu" in cell:
            headers_row = row_idx
            col_map["ten"] = col_idx
        elif "thời lượng" in cell or "thoi luong" in cell or "phút" in cell:
            col_map["thoi_gian_phut"] = col_idx
        elif "danh mục" in cell or "danh muc" in cell:
            col_map["danh_muc"] = col_idx
        elif "số tiền" in cell or "so tien" in cell or "giá" in cell or "gia" in cell:
            col_map["gia_co_ban"] = col_idx
    if headers_row:
        break

# Fallback: cột A=ten, B=thoi_gian, C=danh_muc, D=gia
if not headers_row:
    print("   ⚠️  Không tìm được header, dùng mặc định A=Tên, B=Phút, C=Danh mục, D=Giá")
    headers_row = 1
    col_map = {"ten": 1, "thoi_gian_phut": 2, "danh_muc": 3, "gia_co_ban": 4}

print(f"\n📋 Header row: {headers_row}")
print(f"   Mapping cột: {col_map}")

# ── Parse data ────────────────────────────────────────────────────────────────
def get_cell(row, col_1based):
    """Lấy giá trị cell theo index 1-based"""
    idx = col_1based - 1
    if idx < len(row):
        return row[idx]
    return None

def parse_int(val):
    """Convert giá trị sang integer VNĐ"""
    if val is None:
        return 0
    try:
        f = float(str(val).replace(",", "").replace(".", "").strip())
        # Nếu có dấu thập phân trong số gốc, nhân lại
        s = str(val).strip()
        if "." in s:
            f = float(s)
            if f < 10000:  # Có thể là dạng 1500000.00 đã parse sai
                f = f * 1  # giữ nguyên
        return int(f)
    except:
        return 0

def parse_minutes(val):
    """Convert thời lượng sang integer phút"""
    if val is None:
        return 0
    try:
        return int(float(str(val).strip()))
    except:
        return 0

services = []
thu_tu = 1

for row in sheet.iter_rows(min_row=headers_row + 1, values_only=True):
    ten = get_cell(row, col_map.get("ten", 1))
    if not ten or str(ten).strip() == "":
        continue

    ten = str(ten).strip()
    thoi_gian = parse_minutes(get_cell(row, col_map.get("thoi_gian_phut", 2)))
    danh_muc = get_cell(row, col_map.get("danh_muc", 3))
    danh_muc = str(danh_muc).strip() if danh_muc else None
    if danh_muc in ("None", "", "nan"):
        danh_muc = None

    gia_raw = get_cell(row, col_map.get("gia_co_ban", 4))
    gia = parse_int(gia_raw)

    services.append({
        "ten": ten,
        "thoi_gian_phut": thoi_gian,
        "danh_muc": danh_muc,
        "gia_co_ban": gia,
        "mo_ta": "",
        "mo_ta_ngan": "",
        "ti_le_hoa_hong": 0,
        "thu_tu": thu_tu,
        "hien_tren_menu": True,
        "la_hot": False,
        "is_active": True,
    })
    thu_tu += 1

print(f"\n📊 Đọc được {len(services)} dịch vụ từ Excel")
print("\nMẫu 5 dịch vụ đầu:")
for s in services[:5]:
    print(f"   {s['thu_tu']:3}. {s['ten'][:40]:<40} | {s['thoi_gian_phut']:3}p | {s['danh_muc'] or '—':<25} | {s['gia_co_ban']:>10,}đ")

# ── Xác nhận ──────────────────────────────────────────────────────────────────
print(f"\n⚠️  Sắp insert {len(services)} dịch vụ vào bảng dich_vu")
ans = input("Tiếp tục? (y/n): ").strip().lower()
if ans != "y":
    print("Đã hủy.")
    exit(0)

# ── Xóa data cũ (optional) ────────────────────────────────────────────────────
ans2 = input("Xóa toàn bộ dịch vụ cũ trước khi import? (y/n): ").strip().lower()
if ans2 == "y":
    r = requests.delete(
        f"{SUPABASE_URL}/rest/v1/dich_vu?id=neq.00000000-0000-0000-0000-000000000000",
        headers=HEADERS,
    )
    if r.status_code in (200, 204):
        print("   ✅ Đã xóa dịch vụ cũ")
    else:
        print(f"   ⚠️  Xóa cũ status: {r.status_code} — {r.text[:200]}")

# ── Insert theo batch 50 ──────────────────────────────────────────────────────
BATCH = 50
ok = 0
fail = 0

for i in range(0, len(services), BATCH):
    batch = services[i:i+BATCH]
    r = requests.post(
        f"{SUPABASE_URL}/rest/v1/dich_vu",
        headers=HEADERS,
        data=json.dumps(batch),
    )
    if r.status_code in (200, 201):
        ok += len(batch)
        print(f"   ✅ Batch {i//BATCH + 1}: insert {len(batch)} dịch vụ OK")
    else:
        fail += len(batch)
        print(f"   ❌ Batch {i//BATCH + 1}: lỗi {r.status_code} — {r.text[:300]}")

print(f"\n{'='*50}")
print(f"✅ Thành công: {ok} dịch vụ")
if fail:
    print(f"❌ Thất bại:  {fail} dịch vụ")
print("Done!")
