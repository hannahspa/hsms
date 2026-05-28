"""Import only missing MySpa customers/cards from the legacy exports.

This is intentionally narrow:
- Do not update existing HSMS customers/cards.
- Create placeholder customer phones only when the MySpa phone is missing or
  already belongs to another customer, so old cards can keep their customer id.
- Import only treatment cards created up to 2026-04-30.
"""
import io
import re
import sys
import time
from datetime import datetime
from pathlib import Path

import openpyxl
import requests

sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8")

ROOT = Path(r"C:\Users\Quoc Nam\Projects\hsms")
DB = Path(r"D:\Hannah Spa\Database")
SUPABASE_URL = "https://aqyemkfbjqxpegingoil.supabase.co"
CUTOFF = "2026-04-30"

CUSTOMERS_FILE = DB / "khach_hang_tat_ca_chi_nhanh_1778309550.xlsx"
CARDS_FILE = DB / "danh_sach_the_lieu_trinh_tat_ca_chi_nhanh_1778310007.xlsx"


def load_service_key():
    text = (ROOT / "scripts/import_myspa.py").read_text(encoding="utf-8", errors="ignore")
    match = re.search(r'SERVICE_KEY\s*=\s*"([^"]+)"', text)
    if not match:
        raise RuntimeError("Cannot find SERVICE_KEY")
    return match.group(1)


SERVICE_KEY = load_service_key()
HEADERS = {"apikey": SERVICE_KEY, "Authorization": "Bearer " + SERVICE_KEY}
JSON_HEADERS = {**HEADERS, "Content-Type": "application/json"}


def clean_phone(value):
    text = str(value or "").strip().replace(" ", "").replace(".", "")
    if not text or text.lower() == "none":
        return None
    if text.startswith("+84"):
        text = "0" + text[3:]
    elif text and text[0].isdigit() and not text.startswith("0"):
        text = "0" + text
    return text[:20]


def parse_money(value):
    if value is None:
        return 0
    if isinstance(value, (int, float)):
        return int(value)
    text = str(value).replace(",", "").replace(".00", "").replace("đ", "").strip()
    if not text or text.lower() == "none" or "không" in text.lower():
        return 0
    try:
        return int(float(text))
    except ValueError:
        return 0


def parse_date(value):
    if isinstance(value, datetime):
        return value.date().isoformat()
    if not value:
        return None
    raw = str(value).strip()
    if not raw or raw.lower() == "none" or "không" in raw.lower():
        return None
    for fmt in ("%d/%m/%Y %H:%M:%S", "%d/%m/%Y %H:%M", "%d/%m/%Y", "%Y-%m-%d %H:%M:%S", "%Y-%m-%d"):
        try:
            text = raw[:19] if "%H" in fmt else raw[:10]
            return datetime.strptime(text, fmt).date().isoformat()
        except ValueError:
            pass
    return None


def fetch_all(table, select):
    rows = []
    start = 0
    step = 1000
    while True:
        res = requests.get(
            f"{SUPABASE_URL}/rest/v1/{table}?select={select}",
            headers={**HEADERS, "Range": f"{start}-{start + step - 1}"},
            timeout=120,
        )
        res.raise_for_status()
        batch = res.json()
        rows.extend(batch)
        if len(batch) < step:
            return rows
        start += step


def read_rows(path):
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb.active
    headers = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
    for row in ws.iter_rows(min_row=2, values_only=True):
        if any(v not in (None, "") for v in row):
            yield {headers[i]: row[i] if i < len(row) else None for i in range(len(headers))}
    wb.close()


def load_source_customers():
    customers = {}
    for row in read_rows(CUSTOMERS_FILE):
        ma_kh = str(row.get("Mã khách hàng") or "").strip()
        if not ma_kh:
            continue
        address_parts = [
            str(row.get("Địa chỉ") or "").strip(),
            str(row.get("Phường/Xã") or "").strip(),
            str(row.get("Quận/Huyện") or "").strip(),
            str(row.get("Tỉnh thành") or "").strip(),
        ]
        address = ", ".join([p for p in address_parts if p and p.lower() != "none"]) or None
        gender_raw = str(row.get("Giới tính") or "").strip().lower()
        customers[ma_kh] = {
            "ma_kh": ma_kh,
            "ho_ten": str(row.get("Họ tên") or ma_kh).strip() or ma_kh,
            "so_dien_thoai": clean_phone(row.get("Số điện thoại")),
            "ngay_sinh": parse_date(row.get("Ngày sinh")),
            "gioi_tinh": "nu" if "nữ" in gender_raw else ("nam" if "nam" in gender_raw else "khac"),
            "dia_chi": address,
            "nguon": str(row.get("Nguồn KH") or "MySpa import").strip() or "MySpa import",
            "ghi_chu_da_lieu": str(row.get("Thông tin bệnh lý") or row.get("Ghi chú") or "").strip() or None,
            "tong_chi_tieu": parse_money(row.get("Tổng tiền")),
            "is_active": True,
        }
    return customers


def post_one(table, payload):
    for attempt in range(1, 6):
        res = requests.post(
            f"{SUPABASE_URL}/rest/v1/{table}",
            headers={**JSON_HEADERS, "Prefer": "return=representation"},
            json=payload,
            timeout=120,
        )
        if res.ok:
            data = res.json()
            return data[0] if data else None
        last = f"{res.status_code} {res.text[:300]}"
        time.sleep(min(8, attempt * 1.5))
    raise RuntimeError(f"Cannot insert {table}: {last}")


def main():
    source_customers = load_source_customers()
    db_customers = fetch_all("khach_hang", "id,ma_kh,so_dien_thoai")
    customer_by_code = {r.get("ma_kh"): r for r in db_customers if r.get("ma_kh")}
    used_phones = {r.get("so_dien_thoai") for r in db_customers if r.get("so_dien_thoai")}

    source_cards = []
    for row in read_rows(CARDS_FILE):
        ma_the = str(row.get("Mã Thẻ (Tự động được tạo bởi hệ thống)") or "").strip()
        ngay_mua = parse_date(row.get("Ngày tạo"))
        if not ma_the or not ngay_mua or ngay_mua > CUTOFF:
            continue
        source_cards.append(row)

    db_card_rows = fetch_all("the_lieu_trinh", "ma_the,ngay_mua")
    db_cards = {r.get("ma_the") for r in db_card_rows if r.get("ma_the")}
    db_legacy_cards = {
        r.get("ma_the")
        for r in db_card_rows
        if r.get("ma_the") and (r.get("ngay_mua") or "9999-99-99") <= CUTOFF
    }
    missing_card_rows = [
        row for row in source_cards
        if str(row.get("Mã Thẻ (Tự động được tạo bởi hệ thống)") or "").strip() not in db_legacy_cards
    ]

    needed_customer_codes = {
        str(row.get("Mã khách hàng") or "").strip()
        for row in missing_card_rows
        if str(row.get("Mã khách hàng") or "").strip()
    }
    needed_customer_codes |= set(source_customers) - set(customer_by_code)

    created_customers = 0
    for ma_kh in sorted(needed_customer_codes):
        if ma_kh in customer_by_code:
            continue
        payload = dict(source_customers.get(ma_kh) or {"ma_kh": ma_kh, "ho_ten": ma_kh, "is_active": True})
        phone = payload.get("so_dien_thoai")
        if not phone or phone in used_phones:
            original = phone
            phone = f"MYSPA-{ma_kh}"
            note = payload.get("ghi_chu_da_lieu") or ""
            payload["ghi_chu_da_lieu"] = (note + f"\nMySpa phone gốc: {original or 'trống'}").strip()
        payload["so_dien_thoai"] = phone
        inserted = post_one("khach_hang", payload)
        if inserted:
            customer_by_code[ma_kh] = inserted
            used_phones.add(phone)
            created_customers += 1

    created_cards = 0
    repaired_cards = 0
    skipped_cards = []
    for row in missing_card_rows:
        ma_the = str(row.get("Mã Thẻ (Tự động được tạo bởi hệ thống)") or "").strip()
        ma_kh = str(row.get("Mã khách hàng") or "").strip()
        kh = customer_by_code.get(ma_kh)
        if not kh:
            skipped_cards.append(ma_the)
            continue
        total_raw = str(row.get("Tổng số lần") or "").strip()
        is_unlimited = "không" in total_raw.lower()
        total = parse_money(total_raw)
        used = parse_money(row.get("Số lần đã sử dụng"))
        promo = parse_money(row.get("Số lần khuyến mãi"))
        total_with_promo = total + promo
        if not is_unlimited and total_with_promo <= 0:
            total_with_promo = max(used, 1)
        remaining = 999999 if is_unlimited else max(0, total_with_promo - used)
        expired_at = parse_date(row.get("Ngày hết hạn"))
        payload = {
            "ma_the": ma_the,
            "khach_hang_id": kh["id"],
            "ten_dich_vu": str(row.get("Tên dịch vụ / Tên combo") or row.get("Tên gói") or "").strip() or ma_the,
            "so_buoi_tong": total_with_promo if not is_unlimited else max(used, 0),
            "so_buoi_da_dung": used,
            "gia_tri_the": parse_money(row.get("Tổng tiền")),
            "ngay_mua": parse_date(row.get("Ngày tạo")),
            "ngay_het_han": expired_at,
            "trang_thai": "active" if is_unlimited or remaining > 0 else "het_buoi",
            "loai_the": "combo_lieu_trinh" if str(row.get("Tên dịch vụ / Tên combo") or "").lower().find("combo") >= 0 else "lieu_trinh",
            "is_khong_gioi_han": is_unlimited,
            "source": "myspa_missing_import",
            "ghi_chu": str(row.get("Ghi chú") or "").strip() or None,
            "meta": {
                "myspa_ma_kh": ma_kh,
                "myspa_ten_kh": str(row.get("Tên KH") or "").strip(),
                "myspa_phone": str(row.get("Điện thoại") or "").strip(),
                "myspa_item_code": str(row.get("Mã dịch vụ / Mã combo") or "").strip(),
                "myspa_payment": row.get("Thanh toán"),
                "myspa_debt": row.get("Công nợ"),
                "import_reason": "missing_from_hsms_audit_20260430",
            },
        }
        if ma_the in db_cards:
            patch_payload = dict(payload)
            patch_payload.pop("ma_the", None)
            res = requests.patch(
                f"{SUPABASE_URL}/rest/v1/the_lieu_trinh?ma_the=eq.{requests.utils.quote(ma_the, safe='')}",
                headers=JSON_HEADERS,
                json=patch_payload,
                timeout=120,
            )
            if not res.ok:
                raise RuntimeError(f"Cannot repair card {ma_the}: {res.status_code} {res.text[:300]}")
            repaired_cards += 1
        else:
            post_one("the_lieu_trinh", payload)
            created_cards += 1

    print(f"Created customers: {created_customers}")
    print(f"Created cards: {created_cards}")
    print(f"Repaired existing cards: {repaired_cards}")
    if skipped_cards:
        print("Skipped cards without customer:", ", ".join(skipped_cards))


if __name__ == "__main__":
    main()
