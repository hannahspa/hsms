"""Kiem toan 3 chieu tung ngay: HSMS vs Bank vs MySpa"""
import sys
sys.stdout.reconfigure(encoding='utf-8')
import openpyxl
from supabase import create_client
from collections import defaultdict
from datetime import datetime, timedelta

SUPABASE_URL = 'https://aqyemkfbjqxpegingoil.supabase.co'
SUPABASE_KEY = 'eyJhbGciOiJIUzI1NiIsInR5cCI6IkpXVCJ9.eyJpc3MiOiJzdXBhYmFzZSIsInJlZiI6ImFxeWVta2ZianF4cGVnaW5nb2lsIiwicm9sZSI6InNlcnZpY2Vfcm9sZSIsImlhdCI6MTc3NzUxNTYwMCwiZXhwIjoyMDkzMDkxNjAwfQ.L2yo4Osu6XNhPaOTEMz1Z2GI-SVtzR6AnODirhUR4zI'
supabase = create_client(SUPABASE_URL, SUPABASE_KEY)

def fmt(n):
    return f'{int(n):,}'.replace(',', '') if n else '0'

base = r'D:\Hannah Spa\Thu Chi\Kiem Toan'

# ============================================================
# 1. MYSPA DATA by day + PTTT
# ============================================================
wb = openpyxl.load_workbook(base + '\\DoanhThuMySpa.xlsx', data_only=True)
ws = wb['Sheet1']
myspa = defaultdict(lambda: defaultdict(int))
for row in ws.iter_rows(min_row=2, max_row=ws.max_row, values_only=True):
    ngay_gio = row[6]; pttt = row[9]; so_tien = row[10]
    if so_tien is None or so_tien == 0: continue
    try:
        dt = datetime.strptime(str(ngay_gio).strip().split(' ')[0], '%d/%m/%Y')
        day = dt.strftime('%Y-%m-%d')
    except: continue
    pttt_str = str(pttt).strip() if pttt else 'Khac'
    myspa[day][pttt_str] += int(so_tien)

# ============================================================
# 2. HSMS DATA by day + hinh_thuc
# ============================================================
r = supabase.from_('doanh_thu').select('ngay,hinh_thuc,so_tien').order('ngay').execute()
hsms = defaultdict(lambda: defaultdict(int))
for d in (r.data or []):
    hsms[d['ngay']][d['hinh_thuc']] += (d['so_tien'] or 0)

# ============================================================
# 3. BANK DATA by day - separate QN and KD
# ============================================================
def parse_mb(fp):
    wb2 = openpyxl.load_workbook(fp, data_only=True)
    ws2 = wb2['Lịch sử giao dịch']
    txns = []
    for r in range(20, ws2.max_row + 1):
        date_val = ws2.cell(row=r, column=5).value
        debit = ws2.cell(row=r, column=10).value
        credit = ws2.cell(row=r, column=11).value
        desc = ws2.cell(row=r, column=12).value
        if date_val is None: continue
        try:
            dt = datetime.strptime(str(date_val).strip().split(' ')[0], '%d/%m/%Y')
            day = dt.strftime('%Y-%m-%d')
        except: continue
        no_val = int(float(str(debit).replace(',',''))) if debit else 0
        co_val = int(float(str(credit).replace(',',''))) if credit else 0
        nd = str(desc).strip() if desc else ''
        txns.append({'date': day, 'no': no_val, 'co': co_val, 'desc': nd})
    return txns

kd = parse_mb(base + '\\KhanhDuy.xlsx')
qn = parse_mb(base + '\\QuocNam.xlsx')

def is_ck(t):
    nd = t['desc'].lower()
    if t['co'] == 0: return False
    if any(kw in nd for kw in ['tip em','hoan lai','hoan tra','tra lai','dt tm','tm dt','nop tien mat','tien mat dt','tiền mặt dt','quy','quỹ','lai']):
        return False
    return True

# Bank CK by day, separated by account
bank_qn = defaultdict(int)
bank_kd = defaultdict(int)
for t in qn:
    if is_ck(t): bank_qn[t['date']] += t['co']
for t in kd:
    if is_ck(t): bank_kd[t['date']] += t['co']

# ============================================================
# 4. COMPARE DAY BY DAY - 3 SOURCES
# ============================================================
print('DOI CHIEU 3 CHIEU: MySpa vs HSMS vs Bank (Khach CK)')
print('=' * 100)
print(f'  {"Ngay":<12s} {"MySpa CK":>12s} {"HSMS CK":>12s} {"Bank CK":>12s} {"M-H":>10s} {"M-B":>10s} {"H-B":>10s}  Nguon')

start = '2025-11-26'
end = '2026-05-09'
d = datetime.strptime(start, '%Y-%m-%d')
end_dt = datetime.strptime(end, '%Y-%m-%d')

total_ms = 0; total_hs = 0; total_bk = 0
ok_3 = 0; lech_3 = 0
rows = []

while d <= end_dt:
    day = d.strftime('%Y-%m-%d')
    ms_ck = myspa[day].get('Khách Hàng Chuyển Khoản', 0)
    hs_ck = hsms[day].get('chuyen_khoan', 0)

    # Bank source depends on date
    if day <= '2026-01-07':
        bk_ck = bank_qn.get(day, 0)
        nguon = 'QN'
    else:
        bk_ck = bank_kd.get(day, 0)
        nguon = 'KD'

    total_ms += ms_ck; total_hs += hs_ck; total_bk += bk_ck

    if ms_ck > 0 or hs_ck > 0 or bk_ck > 0:
        mh = ms_ck - hs_ck
        mb = ms_ck - bk_ck
        hb = hs_ck - bk_ck
        if mh == 0 and mb == 0:
            ok_3 += 1
        else:
            lech_3 += 1
        rows.append((day, ms_ck, hs_ck, bk_ck, mh, mb, hb, nguon))
    d += timedelta(days=1)

# Print all rows
for row in rows:
    day, ms, hs, bk, mh, mb, hb, nguon = row
    mh_s = 'OK' if mh == 0 else fmt(mh)
    mb_s = 'OK' if mb == 0 else fmt(mb)
    hb_s = 'OK' if hb == 0 else fmt(hb)
    print(f'  {day:<12s} {fmt(ms):>12s} {fmt(hs):>12s} {fmt(bk):>12s} {mh_s:>10s} {mb_s:>10s} {hb_s:>10s}  {nguon}')

print(f'  {"-"*12} {"-"*12} {"-"*12} {"-"*12}')
print(f'  {"TONG":<12s} {fmt(total_ms):>12s} {fmt(total_hs):>12s} {fmt(total_bk):>12s} {fmt(total_ms-total_hs):>10s} {fmt(total_ms-total_bk):>10s} {fmt(total_hs-total_bk):>10s}')

print(f'\n  OK 3 chieu: {ok_3} | LECH: {lech_3}')
print(f'  MySpa-HSMS: {fmt(total_ms - total_hs)} (M-H)')
print(f'  MySpa-Bank: {fmt(total_ms - total_bk)} (M-B)')
print(f'  HSMS-Bank:  {fmt(total_hs - total_bk)} (H-B)')
