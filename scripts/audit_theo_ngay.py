import os
"""Kiem toan theo tung ngay — tach 2 giai doan Quoc Nam + Khanh Duy"""
import sys
sys.stdout.reconfigure(encoding='utf-8')
import openpyxl
from supabase import create_client
from collections import defaultdict
from datetime import datetime, timedelta

SUPABASE_URL = 'https://aqyemkfbjqxpegingoil.supabase.co'
SUPABASE_KEY = os.environ["SUPABASE_KEY"]
supabase = create_client(SUPABASE_URL, SUPABASE_KEY)

def fmt(n):
    return f'{int(n):,}'.replace(',', '') if n else '0'

base = r'D:\Hannah Spa\Thu Chi\Kiem Toan'

def parse_mb(fp):
    wb = openpyxl.load_workbook(fp, data_only=True)
    ws = wb['Lịch sử giao dịch']
    txns = []
    for r in range(20, ws.max_row + 1):
        date_val = ws.cell(row=r, column=5).value
        debit = ws.cell(row=r, column=10).value
        credit = ws.cell(row=r, column=11).value
        desc = ws.cell(row=r, column=12).value
        if date_val is None: continue
        try:
            dt = datetime.strptime(str(date_val).strip().split(' ')[0], '%d/%m/%Y')
            date_iso = dt.strftime('%Y-%m-%d')
        except: continue
        no_val = int(float(str(debit).replace(',',''))) if debit else 0
        co_val = int(float(str(credit).replace(',',''))) if credit else 0
        nd = str(desc).strip() if desc else ''
        txns.append({'date': date_iso, 'no': no_val, 'co': co_val, 'desc': nd})
    return txns

kd = parse_mb(base + '\\KhanhDuy.xlsx')
qn = parse_mb(base + '\\QuocNam.xlsx')

# HSMS data
r_dt = supabase.from_('doanh_thu').select('ngay,hinh_thuc,so_tien').order('ngay').execute()
r_cp = supabase.from_('chi_phi').select('ngay,hinh_thuc_thanh_toan,so_tien').order('ngay').execute()

hsms_dt_ck = defaultdict(int)
hsms_dt_tm = defaultdict(int)
for d in (r_dt.data or []):
    if d['hinh_thuc'] == 'chuyen_khoan': hsms_dt_ck[d['ngay']] += (d['so_tien'] or 0)
    if d['hinh_thuc'] == 'tien_mat': hsms_dt_tm[d['ngay']] += (d['so_tien'] or 0)

hsms_cp_ck = defaultdict(int)
hsms_cp_tm = defaultdict(int)
for d in (r_cp.data or []):
    if d['hinh_thuc_thanh_toan'] == 'chuyen_khoan': hsms_cp_ck[d['ngay']] += (d['so_tien'] or 0)
    if d['hinh_thuc_thanh_toan'] == 'tien_mat': hsms_cp_tm[d['ngay']] += (d['so_tien'] or 0)

# Bank data grouped by day
def is_customer_transfer(t):
    nd = t['desc'].lower()
    if t['co'] == 0: return False
    if any(kw in nd for kw in ['tip em', 'hoan lai', 'hoan tra', 'tra lai', 'dt tm', 'tm dt', 'nop tien mat', 'nộp tiền mặt', 'tien mat dt', 'tiền mặt dt', 'quy', 'quỹ', 'lai']):
        return False
    return True

bank_ck_by_day = defaultdict(int)
bank_out_by_day = defaultdict(int)
for t in kd + qn:
    if is_customer_transfer(t):
        bank_ck_by_day[t['date']] += t['co']
    if t['no'] > 0:
        bank_out_by_day[t['date']] += t['no']

# ============================================================
# GIAI DOAN 1: QUOC NAM (26/11/2025 - 19/01/2026)
# ============================================================
print('=' * 90)
print('GIAI DOAN 1: TK QUOC NAM (26/11/2025 - 19/01/2026)')
print('=' * 90)
print(f'  {"Ngay":<12s} {"Bank CK vao":>14s} {"HSMS CK":>14s} {"Chenh":>12s}')

start = '2025-11-26'
end = '2026-01-19'
running_bank = 0
running_hsms = 0
total_bank = 0
total_hsms = 0
mismatches = []

d = datetime.strptime(start, '%Y-%m-%d')
end_dt = datetime.strptime(end, '%Y-%m-%d')
while d <= end_dt:
    day = d.strftime('%Y-%m-%d')
    bank = bank_ck_by_day.get(day, 0)
    hsms = hsms_dt_ck.get(day, 0)
    diff = bank - hsms
    total_bank += bank
    total_hsms += hsms
    running_bank += bank
    running_hsms += hsms
    if bank > 0 or hsms > 0:
        icon = 'OK' if diff == 0 else f'LECH {fmt(diff)}'
        print(f'  {day:<12s} {fmt(bank):>14s} {fmt(hsms):>14s} {icon:>12s}')
        if diff != 0: mismatches.append((day, bank, hsms, diff))
    d += timedelta(days=1)

print(f'  {"-"*12} {"-"*14} {"-"*14}')
print(f'  {"TONG":<12s} {fmt(total_bank):>14s} {fmt(total_hsms):>14s} {fmt(total_bank - total_hsms):>12s}')

# ============================================================
# GIAI DOAN 2: KHANH DUY (08/01/2026 - 09/05/2026)
# ============================================================
print()
print('=' * 90)
print('GIAI DOAN 2: TK KHANH DUY (08/01/2026 - 09/05/2026)')
print('=' * 90)
print(f'  {"Ngay":<12s} {"Bank CK vao":>14s} {"HSMS CK":>14s} {"Chenh":>12s}')

start2 = '2026-01-08'
end2 = '2026-05-09'
running_bank2 = 0
running_hsms2 = 0
total_bank2 = 0
total_hsms2 = 0
mismatches2 = []

d2 = datetime.strptime(start2, '%Y-%m-%d')
end_dt2 = datetime.strptime(end2, '%Y-%m-%d')
while d2 <= end_dt2:
    day = d2.strftime('%Y-%m-%d')
    bank = bank_ck_by_day.get(day, 0)
    hsms = hsms_dt_ck.get(day, 0)
    diff = bank - hsms
    total_bank2 += bank
    total_hsms2 += hsms
    running_bank2 += bank
    running_hsms2 += hsms
    if bank > 0 or hsms > 0:
        icon = 'OK' if diff == 0 else f'LECH {fmt(diff)}'
        print(f'  {day:<12s} {fmt(bank):>14s} {fmt(hsms):>14s} {icon:>12s}')
        if diff != 0: mismatches2.append((day, bank, hsms, diff))
    d2 += timedelta(days=1)

print(f'  {"-"*12} {"-"*14} {"-"*14}')
print(f'  {"TONG":<12s} {fmt(total_bank2):>14s} {fmt(total_hsms2):>14s} {fmt(total_bank2 - total_hsms2):>12s}')

# ============================================================
# TONG KET
# ============================================================
print()
print('=' * 90)
print('TONG KET 2 GIAI DOAN')
print('=' * 90)
print(f'  QN: Bank={fmt(total_bank)} HSMS={fmt(total_hsms)} Chenh={fmt(total_bank - total_hsms)}')
print(f'  KD: Bank={fmt(total_bank2)} HSMS={fmt(total_hsms2)} Chenh={fmt(total_bank2 - total_hsms2)}')
print(f'  Tong Bank={fmt(total_bank + total_bank2)} HSMS={fmt(total_hsms + total_hsms2)}')
