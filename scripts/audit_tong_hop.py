import os
"""Kiem toan tong hop: Sao ke MB Bank + HSMS"""
import sys
sys.stdout.reconfigure(encoding='utf-8')
import openpyxl
from supabase import create_client
from collections import defaultdict
from datetime import datetime

SUPABASE_URL = 'https://aqyemkfbjqxpegingoil.supabase.co'
SUPABASE_KEY = os.environ["SUPABASE_KEY"]
supabase = create_client(SUPABASE_URL, SUPABASE_KEY)

def fmt(n):
    if n is None: return '0d'
    return f'{int(n):,}d'.replace(',', '.')

def parse_date(val):
    if isinstance(val, datetime): return val.strftime('%Y-%m-%d')
    if isinstance(val, str):
        for f in ['%d/%m/%Y', '%Y-%m-%d']:
            try: return datetime.strptime(val.strip(), f).strftime('%Y-%m-%d')
            except: pass
    return str(val) if val else '?'

# ============================================================
# DOC CA 2 FILE
# ============================================================
def read_khanh_duy(path):
    """Read raw bank statement (Khanh Duy.xlsx format)"""
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb['Chi Tiết Giao Dịch']
    txns = []
    for row in ws.iter_rows(min_row=3, max_row=ws.max_row, values_only=True):
        stt, ngay, so_bt, no_val, co_val, loai, nd = row
        if stt is None or (no_val is None and co_val is None):
            continue
        if isinstance(stt, str) and 'TỔNG' in stt:
            continue
        date_str = parse_date(ngay)
        txns.append({
            'date': date_str, 'no': int(no_val) if no_val else 0,
            'co': int(co_val) if co_val else 0,
            'loai': str(loai).strip() if loai else '',
            'nd': str(nd).strip() if nd else '',
        })
    return txns

def read_quoc_nam(path):
    """Read classified bank statement (Sao Ke Quoc Nam.xlsx format)"""
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb['Giao Dịch Chi Tiết']
    txns = []
    for row in ws.iter_rows(min_row=4, max_row=ws.max_row, values_only=True):
        ngay, ma_gd, ps_no, ps_co, dien_giai, nhom, loai_ct, ghi_chu = row
        if ngay is None or (ps_no is None and ps_co is None):
            continue
        date_str = parse_date(ngay)
        txns.append({
            'date': date_str, 'no': int(ps_no) if ps_no else 0,
            'co': int(ps_co) if ps_co else 0,
            'nhom': str(nhom).strip() if nhom else '',
            'loai': str(loai_ct).strip() if loai_ct else '',
            'nd': str(dien_giai).strip() if dien_giai else '',
        })
    return txns

base = r'D:\Hannah Spa\Thu Chi\Kiem Toan'
kd_raw = read_khanh_duy(f'{base}\\Sao Ke Khanh Duy.xlsx')
qn_classified = read_quoc_nam(f'{base}\\Sao Ke Quoc Nam.xlsx')

# ============================================================
# 1. TONG QUAN DU LIEU
# ============================================================
print('=' * 80)
print('BAO CAO KIEM TOAN — DOANH THU KHACH CHUYEN KHOAN')
print('=' * 80)

print(f'\nNGUON DU LIEU:')
print(f'  TK MB Bank: 320011919 — CAO QUOC NAM')
print(f'  Ky sao ke: 01/12/2025 – 19/01/2026')
print(f'  File raw (Khanh Duy.xlsx): {len(kd_raw)} giao dich')
print(f'  File classified (Sao Ke Quoc Nam.xlsx): {len(qn_classified)} giao dich')

# Phan loai doanh thu tu ca 2 nguon
kd_dt = [t for t in kd_raw if t['loai'] == 'Doanh Thu' and t['co'] > 0]
qn_dt = [t for t in qn_classified if t['loai'] == 'THU_DV' and t['co'] > 0]

print(f'  Doanh Thu (Khanh Duy raw): {len(kd_dt)} GD, {fmt(sum(t["co"] for t in kd_dt))}')
print(f'  Doanh Thu (Quoc Nam classified): {len(qn_dt)} GD, {fmt(sum(t["co"] for t in qn_dt))}')

# ============================================================
# 2. DOI CHIEU DOANH THU TUNG NGAY (dung Quoc Nam classified)
# ============================================================
print()
print('=' * 80)
print('DOI CHIEU DOANH THU CK — SAO KE vs HSMS (TUNG NGAY)')
print('=' * 80)

sk_by_day = defaultdict(int)
for t in qn_dt:
    sk_by_day[t['date']] += t['co']

r = supabase.from_('doanh_thu').select('ngay,so_tien') \
    .eq('hinh_thuc','chuyen_khoan') \
    .gte('ngay','2025-12-01').lte('ngay','2026-01-31') \
    .order('ngay').execute()

hsms_by_day = defaultdict(int)
for d in (r.data or []):
    hsms_by_day[d['ngay']] += (d['so_tien'] or 0)

all_days = sorted(set(list(sk_by_day.keys()) + list(hsms_by_day.keys())))

mismatches = []
ok_count = 0
thieu_count = 0
thua_count = 0

for day in all_days:
    sk = sk_by_day.get(day, 0)
    hs = hsms_by_day.get(day, 0)
    diff = sk - hs
    if diff == 0:
        ok_count += 1
    elif sk > 0 and hs == 0:
        thieu_count += 1
        mismatches.append(('THIEU', day, sk, hs, diff))
    elif sk == 0 and hs > 0:
        thua_count += 1
        mismatches.append(('THUA', day, sk, hs, diff))
    else:
        mismatches.append(('LECH', day, sk, hs, diff))

# Print summary
print(f'\nTong ngay co GD: {len(all_days)}')
lech_count = len([m for m in mismatches if m[0] == 'LECH'])
print(f'  KHOP:     {ok_count} ngay')
print(f'  LECH:     {lech_count} ngay')
print(f'  THIEU:    {thieu_count} ngay (HSMS khong co nhung sao ke co)')
print(f'  THUA:     {thua_count} ngay (HSMS co nhung sao ke khong co)')

# ============================================================
# 3. THONG KE THEO THANG
# ============================================================
print()
print('=' * 80)
print('DOI CHIEU THEO THANG')
print('=' * 80)

for month in ['2025-12', '2026-01']:
    sk_m = sum(t['co'] for t in qn_dt if t['date'][:7] == month)
    hs_m = sum(d['so_tien'] or 0 for d in (r.data or []) if d['ngay'][:7] == month)
    diff = sk_m - hs_m
    n_sk = len([t for t in qn_dt if t['date'][:7] == month])
    print(f'\n  {month}:')
    print(f'    Sao ke (THU_DV):  {fmt(sk_m)} ({n_sk} GD)')
    print(f'    HSMS (DT CK):     {fmt(hs_m)}')
    print(f'    Chenh lech:       {fmt(diff)}')

# ============================================================
# 4. PHAN TICH GIAI DOAN
# ============================================================
print()
print('=' * 80)
print('PHAN TICH THEO GIAI DOAN')
print('=' * 80)

# Giai doan 1: 01/12-08/01 — TK Quoc Nam nhan tien
gd1_end = '2026-01-08'
sk_gd1 = sum(t['co'] for t in qn_dt if t['date'] <= gd1_end)
hs_gd1 = sum(d['so_tien'] or 0 for d in (r.data or []) if d['ngay'] <= gd1_end)

# Giai doan 2: 09/01-31/01 — TK Khanh Duy nhan tien (sao ke chua co)
sk_gd2 = sum(t['co'] for t in qn_dt if t['date'] > gd1_end)
hs_gd2 = sum(d['so_tien'] or 0 for d in (r.data or []) if d['ngay'] > gd1_end)

print(f'\n  Giai doan 1 (01/12-08/01): TK Quoc Nam')
print(f'    Sao ke:  {fmt(sk_gd1)}')
print(f'    HSMS:    {fmt(hs_gd1)}')
print(f'    Chenh:   {fmt(sk_gd1 - hs_gd1)}')

print(f'\n  Giai doan 2 (09/01-31/01): Chuyen sang TK Khanh Duy')
print(f'    Sao ke:  {fmt(sk_gd2)} (TK Quoc Nam khong con GD)')
print(f'    HSMS:    {fmt(hs_gd2)}')
print(f'    Ghi chu: CAN SAO KE TK KHANH DUY de doi chieu {fmt(hs_gd2)}')

# ============================================================
# 5. CHI TIET SAI LECH LON
# ============================================================
print()
print('=' * 80)
print('TOP 10 NGAY CHENH LECH LON NHAT')
print('=' * 80)

mismatches.sort(key=lambda x: abs(x[4]), reverse=True)
print(f'  {"Ngay":<12s} {"Loai":<8s} {"Sao ke":>14s} {"HSMS":>14s} {"Chenh":>14s}')
for typ, day, sk, hs, diff in mismatches[:15]:
    print(f'  {day:<12s} {typ:<8s} {fmt(sk):>14s} {fmt(hs):>14s} {fmt(diff):>14s}')

# ============================================================
# 6. SO SANH TONG THE
# ============================================================
print()
print('=' * 80)
print('TONG KET KIEM TOAN')
print('=' * 80)

total_sk_dt = sum(t['co'] for t in qn_dt)
total_hs_dt = sum(d['so_tien'] or 0 for d in (r.data or []))

# HSMS toan bo doanh thu CK
r_all = supabase.from_('doanh_thu').select('so_tien').eq('hinh_thuc','chuyen_khoan').execute()
total_hs_all = sum(d['so_tien'] or 0 for d in (r_all.data or []))

print(f'\n  Doanh thu CK trong sao ke (01/12-19/01):  {fmt(total_sk_dt)}')
print(f'  Doanh thu CK trong HSMS (01/12-31/01):    {fmt(total_hs_dt)}')
print(f'  Doanh thu CK trong HSMS (toan bo):         {fmt(total_hs_all)}')
print()
print(f'  NGUON DU LIEU:')
print(f'    - TK Quoc Nam MB Bank: 01/12/2025 – 19/01/2026')
print(f'    - TK Khanh Duy: CHUA CO SAO KE')
print(f'    - T02-T03/2026: CHUA CO SAO KE (ca 2 TK)')
print()
print(f'  KHOANG TRONG DU LIEU:')
print(f'    - 20/01/2026 – 31/03/2026: Khong co sao ke')
print(f'    - 01/04/2026 – 30/04/2026: Chi co 6 ngay tu file cu (25-30/04)')
