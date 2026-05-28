"""Stage MySpa employee commission detail exports into HSMS.

Modes:
  audit  - parse files and compare with HSMS, no writes
  stage  - upsert parsed rows into myspa_commission_detail

Run from repo root:
  python scripts/import_myspa_commission_details.py audit
  python scripts/import_myspa_commission_details.py stage
"""
import io
import json
import re
import sys
import time
import unicodedata
from datetime import datetime
from pathlib import Path

import openpyxl
import requests

sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8")

ROOT = Path(r"C:\Users\Quoc Nam\Projects\hsms")
SOURCE_DIR = Path(r"D:\Hannah Spa\Database\Tien Commission")
SUPABASE_URL = "https://aqyemkfbjqxpegingoil.supabase.co"
BATCH_SIZE = 1000
STAGE_BATCH_SIZE = 250


def load_service_key():
    text = (ROOT / "scripts/import_myspa.py").read_text(encoding="utf-8", errors="ignore")
    match = re.search(r'SERVICE_KEY\s*=\s*"([^"]+)"', text)
    if not match:
        raise RuntimeError("Cannot find SERVICE_KEY in scripts/import_myspa.py")
    return match.group(1)


SERVICE_KEY = load_service_key()
HEADERS = {
    "apikey": SERVICE_KEY,
    "Authorization": "Bearer " + SERVICE_KEY,
}
JSON_HEADERS = {**HEADERS, "Content-Type": "application/json"}


def vn_lower(value):
    text = unicodedata.normalize("NFD", str(value or "").lower())
    text = "".join(ch for ch in text if unicodedata.category(ch) != "Mn")
    text = text.replace("đ", "d")
    return re.sub(r"\s+", " ", text).strip()


def clean_phone(value):
    raw = re.sub(r"\D", "", str(value or ""))
    if not raw:
        return None
    if raw.startswith("84"):
        raw = "0" + raw[2:]
    elif not raw.startswith("0"):
        raw = "0" + raw
    return raw[:15] if len(raw) >= 10 else None


def parse_number(value):
    if value is None or value == "":
        return 0
    if isinstance(value, (int, float)):
        return int(round(float(value)))
    raw = str(value).strip().replace(",", "").replace(".00", "")
    raw = re.sub(r"[^\d\-.]", "", raw)
    try:
        return int(round(float(raw or 0)))
    except ValueError:
        return 0


def parse_decimal(value):
    if value is None or value == "":
        return None
    if isinstance(value, (int, float)):
        return float(value)
    raw = re.sub(r"[^\d\-.]", "", str(value).strip().replace(",", ""))
    try:
        return float(raw) if raw else None
    except ValueError:
        return None


def parse_datetime(value):
    if isinstance(value, datetime):
        return value.isoformat()
    if not value:
        return None
    raw = str(value).strip()
    for fmt in ("%Y-%m-%d %H:%M:%S", "%d/%m/%Y %H:%M:%S", "%d/%m/%Y %H:%M"):
        try:
            return datetime.strptime(raw[:19], fmt).isoformat()
        except ValueError:
            pass
    return None


def date_part(value):
    dt = parse_datetime(value)
    return dt[:10] if dt else None


def short_staff_name(name):
    base = re.sub(r"\(\s*Nghỉ Việc\s*\)", "", str(name or ""), flags=re.IGNORECASE).strip()
    parts = base.split()
    return " ".join(parts[-2:]) if len(parts) > 2 else " ".join(parts)


def staff_name_from_file(path):
    stem = path.stem
    name = re.sub(r"^commission_nhan_vien_", "", stem)
    name = re.sub(r"_tat_ca_chi_nhanh_\d+$", "", name)
    name = name.replace("___", "_").replace("__", "_").strip("_")
    words = [w for w in name.split("_") if w and w not in {"boss"}]
    return " ".join(word.capitalize() for word in words)


def rest_get_all(table, select):
    rows = []
    start = 0
    while True:
        headers = {**HEADERS, "Range": f"{start}-{start + BATCH_SIZE - 1}"}
        response = requests.get(f"{SUPABASE_URL}/rest/v1/{table}?select={select}", headers=headers, timeout=120)
        response.raise_for_status()
        batch = response.json()
        rows.extend(batch)
        if len(batch) < BATCH_SIZE:
            break
        start += BATCH_SIZE
    return rows


def load_staff_map():
    rows = rest_get_all("nhan_vien", "id,ho_ten,trang_thai")
    exact = {}
    last2 = {}
    for row in rows:
        key = vn_lower(row["ho_ten"])
        exact[key] = row
        parts = key.split()
        if len(parts) >= 2:
            last2[" ".join(parts[-2:])] = row
    return exact, last2


def match_staff(name, exact, last2):
    key = vn_lower(name)
    if key in exact:
        return exact[key]
    parts = key.split()
    if len(parts) >= 2:
        return last2.get(" ".join(parts[-2:]))
    return None


def load_order_map():
    rows = rest_get_all("don_hang", "id,ma_don,ngay,is_test")
    return {row["ma_don"]: row for row in rows if row.get("ma_don") and not row.get("is_test")}


def parse_files(exact, last2, order_map):
    parsed = []
    files = sorted(SOURCE_DIR.glob("*.xlsx"))
    for file_path in files:
        staff_raw = staff_name_from_file(file_path)
        matched_staff = match_staff(staff_raw, exact, last2)
        staff_status = "dang_lam" if matched_staff else "nghi_viec"
        staff_display = matched_staff["ho_ten"] if matched_staff else f"{short_staff_name(staff_raw)} (Nghỉ Việc)"

        wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
        ws = wb.active
        headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
        for row_no, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            ma_don_the = str(row[4] or "").strip()
            if not ma_don_the or ma_don_the == "Tổng cộng":
                continue
            order = order_map.get(ma_don_the)
            commission_doanh_thu = parse_number(row[16])
            commission_ngay = parse_number(row[17])
            commission_tong = parse_number(row[18])
            tien_tour = parse_number(row[19])
            tong_tien = commission_ngay + commission_tong + tien_tour
            if tong_tien == 0 and commission_doanh_thu > 0:
                tong_tien = commission_doanh_thu

            parsed.append({
                "source_file": file_path.name,
                "source_row": row_no,
                "staff_file_name": file_path.stem,
                "staff_name_raw": staff_raw,
                "staff_display": staff_display,
                "staff_status": staff_status,
                "matched_nhan_vien_id": matched_staff["id"] if matched_staff else None,
                "thoi_gian_thanh_toan": parse_datetime(row[0]),
                "thoi_gian_them_nhan_vien": parse_datetime(row[1]),
                "thoi_gian_tao_don_hang": parse_datetime(row[2]),
                "ngay": date_part(row[0]) or date_part(row[2]),
                "loai_don_hang": str(row[3] or "").strip() or None,
                "ma_don_the": ma_don_the,
                "ten_dich_vu": str(row[5] or "").strip() or None,
                "item_code": str(row[6] or "").strip() or None,
                "so_luong": parse_decimal(row[7]),
                "thoi_luong": parse_decimal(row[8]),
                "ten_khach_hang": str(row[9] or "").strip() or None,
                "so_dien_thoai": clean_phone(row[10]),
                "chi_nhanh": str(row[11] or "").strip() or None,
                "doanh_so_truoc_giam": parse_number(row[12]),
                "doanh_so_sau_giam": parse_number(row[13]),
                "chuc_vu": str(row[14] or "").strip() or None,
                "ti_le": parse_decimal(row[15]),
                "commission_doanh_thu": commission_doanh_thu,
                "commission_ngay_tim_kiem": commission_ngay,
                "commission_tong_don": commission_tong,
                "tien_tour_nv": tien_tour,
                "tong_tien": tong_tien,
                "row_kind": "detail",
                "don_hang_id": order["id"] if order else None,
                "match_status": "matched_order" if order else "missing_order",
                "raw": {str(headers[i] or f"col_{i}"): row[i] for i in range(min(len(headers), len(row)))},
            })
        wb.close()
    return parsed


def print_audit(rows):
    by_staff = {}
    matched_orders = 0
    missing_orders = 0
    matched_staff = 0
    for row in rows:
        if row["don_hang_id"]:
            matched_orders += 1
        else:
            missing_orders += 1
        if row["matched_nhan_vien_id"]:
            matched_staff += 1
        key = row["staff_display"]
        agg = by_staff.setdefault(key, {"rows": 0, "amount": 0, "tour": 0, "commission": 0})
        agg["rows"] += 1
        agg["amount"] += row["tong_tien"]
        agg["tour"] += row["tien_tour_nv"]
        agg["commission"] += row["commission_ngay_tim_kiem"] + row["commission_tong_don"]

    print(json.dumps({
        "source_dir": str(SOURCE_DIR),
        "files": len(list(SOURCE_DIR.glob("*.xlsx"))),
        "rows": len(rows),
        "matched_staff_rows": matched_staff,
        "retired_or_missing_staff_rows": len(rows) - matched_staff,
        "matched_order_rows": matched_orders,
        "missing_order_rows": missing_orders,
        "total_amount": sum(row["tong_tien"] for row in rows),
        "total_tour": sum(row["tien_tour_nv"] for row in rows),
        "total_commission": sum(row["commission_ngay_tim_kiem"] + row["commission_tong_don"] for row in rows),
        "top_staff": sorted(
            [{"staff": k, **v} for k, v in by_staff.items()],
            key=lambda item: item["amount"],
            reverse=True,
        )[:20],
        "missing_order_sample": [
            {
                "staff": row["staff_display"],
                "ma_don": row["ma_don_the"],
                "ngay": row["ngay"],
                "ten_dich_vu": row["ten_dich_vu"],
                "tong_tien": row["tong_tien"],
            }
            for row in rows if not row["don_hang_id"]
        ][:20],
    }, ensure_ascii=False, indent=2))


def stage_rows(rows):
    response = requests.delete(
        f"{SUPABASE_URL}/rest/v1/myspa_commission_detail?source_file=not.is.null",
        headers=JSON_HEADERS,
        timeout=120,
    )
    if not response.ok:
        raise RuntimeError(f"Cannot clear staging table: {response.status_code} {response.text[:500]}")

    total = len(rows)
    for start in range(0, total, STAGE_BATCH_SIZE):
        batch = rows[start:start + STAGE_BATCH_SIZE]
        last_error = None
        for attempt in range(1, 6):
            try:
                response = requests.post(
                    f"{SUPABASE_URL}/rest/v1/myspa_commission_detail?on_conflict=source_file,source_row",
                    headers={**JSON_HEADERS, "Prefer": "resolution=merge-duplicates,return=minimal"},
                    json=batch,
                    timeout=120,
                )
                if response.ok:
                    last_error = None
                    break
                last_error = RuntimeError(f"{response.status_code} {response.text[:500]}")
            except requests.RequestException as exc:
                last_error = exc
            time.sleep(min(10, attempt * 2))
        if last_error:
            raise RuntimeError(f"Stage failed at row {start}: {last_error}")
        print(f"  staged {min(start + STAGE_BATCH_SIZE, total):,}/{total:,}")


def main():
    mode = sys.argv[1] if len(sys.argv) > 1 else "audit"
    exact, last2 = load_staff_map()
    order_map = load_order_map()
    rows = parse_files(exact, last2, order_map)
    print_audit(rows)
    if mode == "stage":
        stage_rows(rows)
        print("DONE: staged MySpa commission detail rows.")
    elif mode != "audit":
        raise SystemExit("Usage: python scripts/import_myspa_commission_details.py [audit|stage]")


if __name__ == "__main__":
    main()
