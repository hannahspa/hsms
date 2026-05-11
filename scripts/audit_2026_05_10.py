#!/usr/bin/env python3
"""Kiem toan toan dien — 10/05/2026 — FIX PARSE DATE"""
import sys
sys.stdout.reconfigure(encoding='utf-8')
import openpyxl
from supabase import create_client
from collections import defaultdict
from datetime import datetime

SUPABASE_URL = 'https://aqyemkfbjqxpegingoil.supabase.co'
SUPABASE_KEY = 'eyJhbGciOiJIUzI1NiIsInR5cCI6IkpXVCJ9.eyJpc3MiOiJzdXBhYmFzZSIsInJlZiI6ImFxeWVta2ZianF4cGVnaW5nb2lsIiwicm9sZSI6InNlcnZpY2Vfcm9sZSIsImlhdCI6MTc3NzUxNTYwMCwiZXhwIjoyMDkzMDkxNjAwfQ.L2yo4Osu6XNhPaOTEMz1Z2GI-SVtzR6AnODirhUR4zI'
supabase = create_client(SUPABASE_URL, SUPABASE_KEY)

def fmt(n):
    return f'{int(n):,}'.replace(',', '.')

def parse_date_vn(s):
    """Parse date string in various formats"""
    if s is None:
        return None
    if hasattr(s, 'strftime'):
        return s
    if isinstance(s, str):
        for fmt_s in ['%d/%m/%Y %H:%M:%S', '%d/%m/%Y', '%Y-%m-%d']:
            try:
                return datetime.strptime(s.strip(), fmt_s)
            except:
                pass
    return None

base = r'D:\Hannah Spa\Thu Chi\Kiem Toan'

# ============================================================
# 1. DOANH THU MYSPA — PARSE DUNG
# ============================================================
print('=' * 80)
print('1. DOANH THU MYSPA — PHAN TICH THEO PTTT')
print('=' * 80)

wb = openpyxl.load_workbook(base + '\\DoanhThuMySpa.xlsx', data_only=True)
ws = wb['Sheet1']

myspa_pttt = defaultdict(int)
myspa_monthly = defaultdict(int)
myspa_daily = defaultdict(int)
myspa_total = 0
myspa_count = 0
dt_records_raw = []

for row in ws.iter_rows(min_row=2, max_row=ws.max_row, values_only=True):
    ma_tt = row[0]
    ma_dh = row[1]
    ho_ten = row[2]
    sdt = row[3]
    ngay_gio = row[6]
    thu_ngan = row[7]
    pttt = row[9]
    so_tien = row[10]
    nop_kt = row[11]

    if so_tien is None or so_tien == 0:
        continue

    pttt_str = str(pttt).strip() if pttt else 'Khac'
    if 'Tổng' in pttt_str:
        continue
    if pttt_str == 'None':
        pttt_str = 'Khac'

    amt = int(so_tien)
    myspa_pttt[pttt_str] += amt
    myspa_total += amt
    myspa_count += 1

    d = parse_date_vn(ngay_gio)
    if d:
        myspa_monthly[d.strftime('%Y-%m')] += amt
        myspa_daily[d.strftime('%Y-%m-%d')] += amt

    dt_records_raw.append({
        'ma_tt': ma_tt,
        'ma_dh': ma_dh,
        'ho_ten': ho_ten,
        'sdt': sdt,
        'ngay_gio': ngay_gio,
        'thu_ngan': thu_ngan,
        'pttt': pttt_str,
        'so_tien': amt,
        'nop_kt': nop_kt,
    })

print(f'\nTong giao dich: {myspa_count}')
print(f'Tong doanh thu: {fmt(myspa_total)}')

print(f'\n--- THEO PTTT ---')
print(f'{"PTTT":<35s} {"So GD":>6s} {"So tien":>18s} {"%":>7s}')
print('-' * 70)
for k, v in sorted(myspa_pttt.items(), key=lambda x: -x[1]):
    pct = v / myspa_total * 100 if myspa_total else 0
    print(f'{k:<35s} {sum(1 for r in dt_records_raw if r["pttt"]==k):>6d} {fmt(v):>18s} {pct:>6.1f}%')

print(f'\n--- THEO THANG ---')
for m in sorted(myspa_monthly.keys()):
    v = myspa_monthly[m]
    print(f'  {m}: {fmt(v)}')

# ============================================================
# 2. HSMS DOANH THU
# ============================================================
print()
print('=' * 80)
print('2. DOI CHIEU DOANH THU MySpa vs HSMS')
print('=' * 80)

r_dt = supabase.from_('doanh_thu').select('hinh_thuc,so_tien,ngay').execute()
hsms_pttt = defaultdict(int)
hsms_monthly = defaultdict(int)
for d in (r_dt.data or []):
    hsms_pttt[d['hinh_thuc']] += (d['so_tien'] or 0)
    if d.get('ngay'):
        hsms_monthly[d['ngay'][:7]] += (d['so_tien'] or 0)
hsms_total = sum(hsms_pttt.values())

pttt_map = {
    'Tiền mặt': 'tien_mat',
    'Khách Hàng Chuyển Khoản': 'chuyen_khoan',
    'Quẹt thẻ': 'quet_the',
    'Thẻ trả trước': 'the_tra_truoc',
    'Không xác định': 'unknown',
}

print(f'\n{"Hinh thuc":<35s} {"MySpa":>16s} {"HSMS":>16s} {"Chenh":>16s}')
print('-' * 85)
total_ms_map = 0
total_hs_map = 0
for ms_name, hs_key in pttt_map.items():
    ms = myspa_pttt.get(ms_name, 0)
    hs = hsms_pttt.get(hs_key, 0)
    total_ms_map += ms
    total_hs_map += hs
    diff = ms - hs
    print(f'{ms_name:<35s} {fmt(ms):>16s} {fmt(hs):>16s} {fmt(diff):>16s}')

# Show any unmapped MySpa
for ms_name, v in sorted(myspa_pttt.items(), key=lambda x: -x[1]):
    if ms_name not in pttt_map:
        print(f'{ms_name:<35s} {fmt(v):>16s} {"--":>16s} {"MySpa only":>16s}')

print(f'{"-"*35} {"-"*16} {"-"*16} {"-"*16}')
print(f'{"TONG":<35s} {fmt(total_ms_map):>16s} {fmt(total_hs_map):>16s} {fmt(total_ms_map - total_hs_map):>16s}')

# Monthly comparison
print(f'\n--- THEO THANG ---')
all_months = sorted(set(list(myspa_monthly.keys()) + list(hsms_monthly.keys())))
print(f'{"Thang":<10s} {"MySpa":>18s} {"HSMS":>18s} {"Chenh":>18s}')
for m in all_months:
    ms = myspa_monthly.get(m, 0)
    hs = hsms_monthly.get(m, 0)
    diff = ms - hs
    print(f'{m:<10s} {fmt(ms):>18s} {fmt(hs):>18s} {fmt(diff):>18s}')

# ============================================================
# 3. PHIEU CHI MYSPA
# ============================================================
print()
print('=' * 80)
print('3. PHIEU CHI MYSPA')
print('=' * 80)

wb2 = openpyxl.load_workbook(base + '\\PhieuChiMySpa.xlsx', data_only=True)
ws2 = wb2['Sheet1']
headers = [c.value for c in next(ws2.iter_rows(min_row=1, max_row=1))]
print(f'Headers: {headers}')

myspa_cp = defaultdict(int)
myspa_cp_list = []
for row in ws2.iter_rows(min_row=2, max_row=ws2.max_row, values_only=True):
    so_tien = row[6] if len(row) > 6 else 0  # Column G
    pttt = row[7] if len(row) > 7 else ''     # Column H
    dien_giai = row[9] if len(row) > 9 else ''  # Column J (Ly do)
    ngay = row[1] if len(row) > 1 else None    # Column B (Ngay gio)
    phieu = row[2] if len(row) > 2 else None

    if so_tien is None or so_tien == 0:
        continue

    amt = abs(int(so_tien))
    pttt_str = str(pttt).strip() if pttt else ''

    if 'chuyển khoản' in pttt_str.lower() or 'Chuyển khoản' in pttt_str:
        key = 'chuyen_khoan'
    elif 'tiền mặt' in pttt_str.lower() or 'Tiền mặt' in pttt_str:
        key = 'tien_mat'
    elif 'tổng' in pttt_str.lower():
        continue
    else:
        key = 'khac'

    myspa_cp[key] += amt
    myspa_cp_list.append({
        'pttt': pttt_str,
        'key': key,
        'so_tien': amt,
        'dien_giai': str(dien_giai).strip() if dien_giai else '',
        'phieu': str(phieu).strip() if phieu else '',
        'ngay': str(ngay) if ngay else '',
    })

myspa_cp_total = sum(myspa_cp.values())
print(f'\nMySpa Chi Phi:')
for k, v in sorted(myspa_cp.items(), key=lambda x: -x[1]):
    count = sum(1 for c in myspa_cp_list if c['key'] == k)
    print(f'  {k}: {fmt(v)} ({count} phieu)')
print(f'  Tong: {fmt(myspa_cp_total)}')

# Show some expense details
print(f'\nTop 10 chi phi lon nhat CK:')
ck_items = sorted([c for c in myspa_cp_list if c['key']=='chuyen_khoan'], key=lambda x: -x['so_tien'])[:10]
for i, c in enumerate(ck_items):
    print(f'  {i+1}. {fmt(c["so_tien"])} - {c["phieu"]} - {c["dien_giai"][:80]}')

# ============================================================
# 4. HSMS CHI PHI
# ============================================================
print()
print('=' * 80)
print('4. DOI CHIEU CHI PHI MySpa vs HSMS')
print('=' * 80)

r_cp = supabase.from_('chi_phi').select('hinh_thuc_thanh_toan,so_tien,ngay').execute()
hsms_cp_pttt = defaultdict(int)
hsms_cp_monthly = defaultdict(int)
for d in (r_cp.data or []):
    hsms_cp_pttt[d['hinh_thuc_thanh_toan']] += (d['so_tien'] or 0)
    if d.get('ngay'):
        hsms_cp_monthly[d['ngay'][:7]] += (d['so_tien'] or 0)
hsms_cp_total = sum(hsms_cp_pttt.values())

print(f'\n{"PTTT":<20s} {"MySpa":>16s} {"HSMS":>16s} {"Chenh":>16s}')
print('-' * 68)
for key, label in [('chuyen_khoan', 'CK'), ('tien_mat', 'Tien mat')]:
    ms = myspa_cp.get(key, 0)
    hs = hsms_cp_pttt.get(key, 0)
    diff = ms - hs
    print(f'{label:<20s} {fmt(ms):>16s} {fmt(hs):>16s} {fmt(diff):>16s}')

print(f'{"-"*20} {"-"*16} {"-"*16} {"-"*16}')
print(f'{"TONG CP":<20s} {fmt(myspa_cp_total):>16s} {fmt(hsms_cp_total):>16s} {fmt(myspa_cp_total - hsms_cp_total):>16s}')

# ============================================================
# 5. SO DU VI & CK NOI BO
# ============================================================
print()
print('=' * 80)
print('5. PHAN TICH SO DU VI — HSMS')
print('=' * 80)

r_vi = supabase.from_('vi').select('*').order('thu_tu').execute()
vi_map = {}
for v in (r_vi.data or []):
    vi_map[v['loai']] = v
    print(f"  {v['ten']:12s} id={v['id'][:8]}... so_du_dau={fmt(v.get('so_du_dau', 0) or 0)}")

# Get CK nội bộ
r_cknb = supabase.from_('chuyen_khoan_noi_bo').select('*').execute()
ck_by_type = defaultdict(int)
for c in (r_cknb.data or []):
    tu = c['tu_vi_id']
    den = c['den_vi_id']
    so_tien = c['so_tien'] or 0

    tm_id = vi_map.get('tien_mat', {}).get('id')
    mb_id = vi_map.get('chuyen_khoan', {}).get('id')
    tp_id = vi_map.get('quet_the', {}).get('id')

    if tu == tm_id and den == mb_id:
        ck_by_type['TM->MB'] += so_tien
    elif tu == tp_id and den == mb_id:
        ck_by_type['TP->MB'] += so_tien
    elif tu == mb_id and den == tm_id:
        ck_by_type['MB->TM'] += so_tien
    elif tu == mb_id and den == tp_id:
        ck_by_type['MB->TP'] += so_tien
    elif tu == tm_id and den == tp_id:
        ck_by_type['TM->TP'] += so_tien
    elif tu == tp_id and den == tm_id:
        ck_by_type['TP->TM'] += so_tien

print(f'\nChuyen khoan noi bo:')
for k, v in sorted(ck_by_type.items(), key=lambda x: -x[1]):
    print(f'  {k}: {fmt(v)}')

# Calculate MB Bank
hsms_dt_ck = hsms_pttt.get('chuyen_khoan', 0)
hsms_dt_qt = hsms_pttt.get('quet_the', 0)
hsms_dt_tm = hsms_pttt.get('tien_mat', 0)
hsms_dt_ttt = hsms_pttt.get('the_tra_truoc', 0)

hsms_cp_ck = hsms_cp_pttt.get('chuyen_khoan', 0)
hsms_cp_tm = hsms_cp_pttt.get('tien_mat', 0)

ck_in_mb = ck_by_type.get('TM->MB', 0) + ck_by_type.get('TP->MB', 0)
ck_out_mb = ck_by_type.get('MB->TM', 0) + ck_by_type.get('MB->TP', 0)

mb_dau = vi_map.get('chuyen_khoan', {}).get('so_du_dau', 0) or 0
tp_dau = vi_map.get('quet_the', {}).get('so_du_dau', 0) or 0
tm_dau = vi_map.get('tien_mat', {}).get('so_du_dau', 0) or 0

mb_hsms = mb_dau + hsms_dt_ck + ck_in_mb - hsms_cp_ck - ck_out_mb
tp_hsms = tp_dau + hsms_dt_qt - ck_by_type.get('TP->MB', 0) + ck_by_type.get('MB->TP', 0) + ck_by_type.get('TM->TP', 0) - ck_by_type.get('TP->TM', 0)

print(f'\n--- SO DU MB BANK (HSMS) ---')
print(f'  So du dau:          {fmt(mb_dau)}')
print(f'  + DT Chuyen khoan:  {fmt(hsms_dt_ck)}')
print(f'  + CK vao MB:        {fmt(ck_in_mb)}')
print(f'  - CP Chuyen khoan:  {fmt(hsms_cp_ck)}')
print(f'  - CK ra MB:         {fmt(ck_out_mb)}')
print(f'  = SO DU MB:         {fmt(mb_hsms)}')

print(f'\n--- SO DU TP BANK (HSMS) ---')
print(f'  So du dau:          {fmt(tp_dau)}')
print(f'  + DT Quet the:      {fmt(hsms_dt_qt)}')
print(f'  - CK TP->MB:        {fmt(ck_by_type.get("TP->MB", 0))}')
print(f'  = SO DU TP:         {fmt(tp_hsms)}')

# Actual from view
r_sd = supabase.from_('so_du_vi_thuc_te').select('ten,so_du_hien_tai').order('thu_tu').execute()
print(f'\n--- SO DU VIEW (so_du_vi_thuc_te) ---')
for v in (r_sd.data or []):
    print(f"  {v['ten']:12s} {fmt(v.get('so_du_hien_tai', 0) or 0)}")

# ============================================================
# 6. SO DU THUC TE vs HSMS
# ============================================================
print()
print('=' * 80)
print('6. DOI CHIEU THUC TE vs HSMS')
print('=' * 80)

# From the bank statements
kd_mb_thuc_te = 73_916_234
tp_thuc_te = 29_059_129
vcb_thuc_te = 25_000_000  # uoc tinh

print(f'\n{"Tai khoan":<25s} {"HSMS":>16s} {"Thuc te":>16s} {"Chenh":>16s}')
print('-' * 75)
print(f'{"MB Bank (Khanh Duy)":<25s} {fmt(mb_hsms):>16s} {fmt(kd_mb_thuc_te):>16s} {fmt(mb_hsms - kd_mb_thuc_te):>16s}')
print(f'{"TP Bank (Diem My)":<25s} {fmt(tp_hsms):>16s} {fmt(tp_thuc_te):>16s} {fmt(tp_hsms - tp_thuc_te):>16s}')
print(f'{"VCB (Diem My)":<25s} {"--":>16s} {fmt(vcb_thuc_te):>16s} {"--":>16s}')
print(f'{"Tong Quet The (TP+VCB)":<25s} {fmt(tp_hsms):>16s} {fmt(tp_thuc_te+vcb_thuc_te):>16s} {fmt(tp_hsms - tp_thuc_te - vcb_thuc_te):>16s}')

# ============================================================
# 7. ANALYZE SAO KE NGAN HANG
# ============================================================
print()
print('=' * 80)
print('7. PHAN TICH SAO KE NGAN HANG')
print('=' * 80)

# 7.1 Quoc Nam MB Bank
print('\n--- QUOC NAM MB Bank (320011919) ---')
wb_qn = openpyxl.load_workbook(base + '\\QuocNam.xlsx', data_only=True)
ws_qn = wb_qn['Lịch sử giao dịch']

total_nhap_qn = 0
total_xuat_qn = 0
transfers_to_kd = 0
salary_payments = 0
other_out = 0
customer_in = 0

for row in ws_qn.iter_rows(min_row=20, max_row=ws_qn.max_row, values_only=True):
    stt = row[1]
    ngay = row[4]
    so_bt = row[6]
    phat_sinh_no = row[9]  # money OUT
    phat_sinh_co = row[10]  # money IN
    noi_dung = str(row[11]) if row[11] else ''

    if stt is None or str(stt).strip() == '':
        continue
    if 'STT' in str(stt):
        continue

    try:
        no_amt = float(str(phat_sinh_no).replace(',', '')) if phat_sinh_no else 0
        co_amt = float(str(phat_sinh_co).replace(',', '')) if phat_sinh_co else 0
    except:
        continue

    total_xuat_qn += no_amt
    total_nhap_qn += co_amt

    if no_amt > 0:
        if 'KHANH DUY' in noi_dung.upper():
            transfers_to_kd += no_amt
        elif any(w in noi_dung.upper() for w in ['LUONG', 'KY QUY', 'TIP EM']):
            salary_payments += no_amt
        else:
            other_out += no_amt

    if co_amt > 0:
        customer_in += co_amt

print(f'  Tien vao (Credit):    {fmt(total_nhap_qn)}')
print(f'  Tien ra (Debit):      {fmt(total_xuat_qn)}')
print(f'    -> Chuyen cho KD:   {fmt(transfers_to_kd)}')
print(f'    -> Luong/Ky quy:    {fmt(salary_payments)}')
print(f'    -> Khac:            {fmt(other_out)}')

# 7.2 Khanh Duy MB Bank
print('\n--- KHANH DUY MB Bank (0379080909) ---')
wb_kd = openpyxl.load_workbook(base + '\\KhanhDuy.xlsx', data_only=True)
ws_kd = wb_kd['Lịch sử giao dịch']

total_in_kd = 0
total_out_kd = 0

for row in ws_kd.iter_rows(min_row=20, max_row=ws_kd.max_row, values_only=True):
    stt = row[1]
    ngay = row[4]
    phat_sinh_no = row[9]
    phat_sinh_co = row[10]

    if stt is None or str(stt).strip() == '':
        continue
    if 'STT' in str(stt):
        continue

    try:
        no_amt = float(str(phat_sinh_no).replace(',', '')) if phat_sinh_no else 0
        co_amt = float(str(phat_sinh_co).replace(',', '')) if phat_sinh_co else 0
    except:
        continue

    total_in_kd += co_amt
    total_out_kd += no_amt

print(f'  Tien vao (Credit):    {fmt(total_in_kd)}')
print(f'  Tien ra (Debit):      {fmt(total_out_kd)}')
print(f'  Net flow:             {fmt(total_in_kd - total_out_kd)}')

# Determine initial balance from the first row
print('\n  Tim so du dau...')
for row in ws_kd.iter_rows(min_row=1, max_row=19, values_only=True):
    txt = ' '.join([str(c) for c in row if c])
    if txt.strip():
        pass  # just print the header info

# 7.3 Sao Ke Quet The (TP Bank)
print('\n--- SAO KE QUET THE (TP Bank - Diem My) ---')
wb_qt = openpyxl.load_workbook(base + '\\Sao Ke Quet The.xlsx', data_only=True)
ws_qt = wb_qt['Sao kê giao dịch']

total_qt_in = 0
total_qt_out = 0
to_khanh_duy = 0
for row in ws_qt.iter_rows(min_row=5, max_row=ws_qt.max_row, values_only=True):
    ngay, nd, so_tien, so_du = row[0], row[1], row[2], row[3]
    if so_tien is None:
        continue
    if isinstance(so_tien, str):
        try:
            so_tien = float(so_tien.replace(',', ''))
        except:
            continue
    so_tien = int(so_tien)
    nd_str = str(nd) if nd else ''

    if so_tien > 0:
        total_qt_in += so_tien
    elif so_tien < 0:
        total_qt_out += abs(so_tien)
        if 'KHANH DUY' in nd_str.upper():
            to_khanh_duy += abs(so_tien)

print(f'  Tong tien vao:     {fmt(total_qt_in)}')
print(f'  Tong tien ra:      {fmt(total_qt_out)}')
print(f'    -> Cho KD:       {fmt(to_khanh_duy)}')
print(f'  Net flow:          {fmt(total_qt_in - total_qt_out)}')
print(f'  So du cuoi:        {fmt(29_059_129)}')  # from the last row

# ============================================================
# 8. TONG KET — GIAI THICH CHENH LECH
# ============================================================
print()
print('=' * 80)
print('8. TONG KET — GIAI THICH NGUYEN NHAN CHENH LECH')
print('=' * 80)

print("""
CAU HOI: Tai sao HSMS/MySpa bao cao:
  - MB Bank: 160.511.484d  nhung thuc te Khanh Duy chi giu: 73.916.234d?
  - TP Bank:  78.554.000d  nhung thuc te TP Bank chi co:  29.059.129d?

CONG THUC TINH SO DU HSMS:
  MB = so_du_dau + DT_CK + CK_vao_MB - CP_CK - CK_ra_MB
  TP = so_du_dau + DT_QT - CK_TP->MB

CAC YEU TO GAY CHENH LECH:
""")

# Key metrics
print(f'1. SO DU DAU = 0 (KHONG CO SO DU BAN DAU)')
print(f'   Neu MB Bank da co tien truoc khi bat dau HSMS, so du dau phai > 0')
print(f'   Nhung dieu nay lam HSMS BAO CAO THAP HON, khong phai cao hon')

print(f'\n2. DOANH THU CK = {fmt(hsms_dt_ck)}')
print(f'   Khach CK truc tiep vao TK nao?')
print(f'   Neu vao TK Quoc Nam (320011919) thay vi TK Khanh Duy (0379080909)')
print(f'   Thi so du MB Bank (Khanh Duy) se khong co tien CK cua KHACH')

print(f'\n3. CHI PHI CK = {fmt(hsms_cp_ck)}')
print(f'   MySpa CP CK = {fmt(myspa_cp.get("chuyen_khoan", 0))}')
print(f'   HSMS CP CK nhieu hon MySpa: {fmt(hsms_cp_ck - myspa_cp.get("chuyen_khoan", 0))}')
print(f'   ---> HSMS CO THE DANG GHI NHAN CHI PHI TU 2 NGUON (MySpa + HSMS)')

print(f'\n4. CHUYEN KHOAN NOI BO VAO MB = {fmt(ck_in_mb)}')
print(f'   TM->MB: {fmt(ck_by_type.get("TM->MB", 0))} (tien mat nop vao NH)')
print(f'   TP->MB: {fmt(ck_by_type.get("TP->MB", 0))} (quet the chuyen ve MB)')

print(f'\n5. QUET THE THUC TE:')
print(f'   HSMS QT:  {fmt(hsms_dt_qt)}')
print(f'   TP Bank:   {fmt(tp_thuc_te)} (so du cuoi)')
print(f'   VCB:      ~{fmt(vcb_thuc_te)} (uoc tinh)')
print(f'   TP In:     {fmt(total_qt_in)} (tong tien vao TP)')
print(f'   TP Out:    {fmt(total_qt_out)} (da chuyen cho KD)')
print(f'   ---> Tien quet the da bi chuyen bot sang Khanh Duy')
print(f'        HSMS khong tru khoan chuyen nay ra khoi so du TP')

# Cashflow analysis
print(f'\n6. DONG TIEN THUC TE:')
print(f'   Khach CK -> TK Quoc Nam ({fmt(customer_in)}) -> 1 phan chuyen cho KD ({fmt(transfers_to_kd)})')
print(f'   Khach Quet the -> TP Bank ({fmt(total_qt_in)}) -> 1 phan chuyen cho KD ({fmt(to_khanh_duy)})')
print(f'   Tong tien vao TK KD tu 2 nguon: {fmt(transfers_to_kd + to_khanh_duy)}')

# Net calculation for KD
print(f'\n7. NGUYEN NHAN CHINH:')
print(f'   a) HSMS tinh DT CK = {fmt(hsms_dt_ck)} dua tren doanh thu bao cao')
print(f'      Nhung KHACH CK VAO TK QUOC NAM, chi 1 phan chuyen cho KD')
print(f'   b) HSMS tinh DT QT = {fmt(hsms_dt_qt)}')
print(f'      Nhung 1 phan da chuyen cho KD ({fmt(to_khanh_duy)}), HSMS khong tru')
print(f'   c) HSMS CP CK = {fmt(hsms_cp_ck)} > MySpa CP CK = {fmt(myspa_cp.get("chuyen_khoan", 0))}')
print(f'      HSMS ghi nhan them {fmt(hsms_cp_ck - myspa_cp.get("chuyen_khoan", 0))} chi phi CK')
