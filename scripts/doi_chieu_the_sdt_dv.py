# -*- coding: utf-8 -*-
"""
ĐỐI CHIẾU THẺ LIỆU TRÌNH MySpa ↔ HSMS — chuẩn bị cắt MySpa (~15/07/2026).

QUY TẮC (chốt 03/07/2026 — xem memory feedback_khong_doi_soat_the_theo_ma):
  ✗ KHÔNG match theo ma_the (2 hệ lệch mã từ ~06/2026, khác cả khách)
  ✓ Match theo bộ khóa:  SĐT khách (chuẩn hóa)  +  tên dịch vụ (chuẩn hóa)
  ✓ Rồi so danh sách SỐ BUỔI CÒN LẠI (multiset 2 bên, đã sort)

CÁCH CHẠY (từ thư mục gốc repo):
  python scripts/doi_chieu_the_sdt_dv.py [đường_dẫn_excel_myspa]
  - Không truyền arg → tự lấy file danh_sach_the_lieu_trinh_tat_ca_chi_nhanh_*.xlsx
    MỚI NHẤT trong D:\\Hannah Spa\\Database
  - Cần .env.import (SUPABASE_URL + SUPABASE_KEY service_role)

KẾT QUẢ: D:\\Hannah Spa\\Database\\DOI_CHIEU_THE_SDT_DV_<ngày>.xlsx — 4 sheet:
  1. KHOP           — nhóm (SĐT, DV) 2 bên khớp buổi còn lại 100%
  2. LECH_BUOI      — cùng khách + DV nhưng buổi còn lại lệch → xem tay
  3. MYSPA_CO_HSMS_THIEU — MySpa có, HSMS không thấy → cần bổ sung trước khi cắt
  4. HSMS_DU        — HSMS có, MySpa không (thẻ bán sau ngày export = bình thường)
CHỈ ĐỌC — không sửa gì trong DB.
"""
import sys, io, os, glob, re, unicodedata
from datetime import datetime
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')

import openpyxl
from openpyxl import Workbook
from supabase import create_client
from dotenv import load_dotenv

load_dotenv('.env.import')
load_dotenv('.env')
URL = os.environ.get('SUPABASE_URL') or os.environ.get('VITE_SUPABASE_URL') or 'https://api.hannahspa.vn'
sb = create_client(URL, os.environ['SUPABASE_KEY'])

DB_DIR = r"D:\Hannah Spa\Database"

# ── Chuẩn hóa ────────────────────────────────────────────────────────────────
def chuan_sdt(s):
    """0907xxx / +84907xxx / 84907xxx → 0907xxx; rác → ''"""
    d = re.sub(r'\D', '', str(s or ''))
    if d.startswith('84') and len(d) > 9:
        d = '0' + d[2:]
    if d and not d.startswith('0'):
        d = '0' + d
    return d if len(d) >= 9 else ''

def bo_dau(s):
    s = unicodedata.normalize('NFD', str(s or ''))
    s = ''.join(c for c in s if unicodedata.category(c) != 'Mn')
    return s.replace('đ', 'd').replace('Đ', 'D')

def chuan_dv(s):
    """Tên dịch vụ → khóa so sánh: bỏ dấu, thường, bỏ ký tự lạ, nén khoảng trắng."""
    s = bo_dau(s).lower()
    s = re.sub(r'[^a-z0-9 ]', ' ', s)
    return re.sub(r'\s+', ' ', s).strip()

def to_int(x):
    try:
        return int(float(str(x).replace(',', '')))
    except (ValueError, TypeError):
        return 0

# ── 1. Đọc MySpa export ──────────────────────────────────────────────────────
if len(sys.argv) > 1:
    excel_path = sys.argv[1]
else:
    files = sorted(glob.glob(os.path.join(DB_DIR, 'danh_sach_the_lieu_trinh_tat_ca_chi_nhanh_*.xlsx')))
    if not files:
        sys.exit(f'Không thấy file export MySpa trong {DB_DIR}')
    excel_path = files[-1]
print(f'📂 MySpa export: {excel_path}')

wb = openpyxl.load_workbook(excel_path, read_only=True, data_only=True)
ws = wb.active
# Cột export MySpa (đã dùng ổn định từ đợt 05/2026):
# 0=mã thẻ 1=ngày tạo 3=tên KH 4=SĐT 5=tên gói 7=tên DV 8=buổi tổng 9=buổi đã dùng
# 11=tổng tiền 13=công nợ 15=hết hạn 17=chi nhánh
myspa_rows = []
for i, row in enumerate(ws.iter_rows(values_only=True)):
    if i == 0 or not row[0]:
        continue
    tong, dung = to_int(row[8]), to_int(row[9])
    myspa_rows.append({
        'ma_the': str(row[0]).strip(), 'ten_kh': row[3], 'sdt': chuan_sdt(row[4]),
        'ten_dv': row[5] or row[7], 'dv_key': chuan_dv(row[5] or row[7]),
        'tong': tong, 'dung': dung, 'con': max(0, tong - dung),
        'cong_no': to_int(row[13]),
    })
print(f'   → {len(myspa_rows)} thẻ MySpa')

# ── 2. Đọc HSMS (phân trang 1000) ────────────────────────────────────────────
hsms_rows, offset = [], 0
while True:
    res = sb.table('the_lieu_trinh').select(
        'id, ma_the, ten_dich_vu, so_buoi_tong, so_buoi_da_dung, so_buoi_con_lai, '
        'trang_thai, da_xoa, bi_dong, '
        'khach_hang:khach_hang_id(ho_ten, so_dien_thoai)'
    ).range(offset, offset + 999).execute()
    batch = res.data or []
    hsms_rows.extend(batch)
    if len(batch) < 1000:
        break
    offset += 1000
# Loại thẻ đã xóa mềm (da_xoa) khỏi so sánh
n_da_xoa = sum(1 for h in hsms_rows if h.get('da_xoa'))
hsms_rows = [h for h in hsms_rows if not h.get('da_xoa')]
for h in hsms_rows:
    kh = h.get('khach_hang') or {}
    h['sdt'] = chuan_sdt(kh.get('so_dien_thoai'))
    h['ten_kh'] = kh.get('ho_ten')
    h['dv_key'] = chuan_dv(h.get('ten_dich_vu'))
    tong = h.get('so_buoi_tong') or 0
    dung = h.get('so_buoi_da_dung') or 0
    h['con'] = max(0, tong - dung)
print(f'   → {len(hsms_rows)} thẻ HSMS (đã loại {n_da_xoa} thẻ xóa mềm)')

# ── 3. Nhóm theo (SĐT, DV) và so multiset buổi còn lại ──────────────────────
def nhom(rows):
    g = {}
    for r in rows:
        if not r['sdt']:
            continue  # không SĐT → không match tự động được
        g.setdefault((r['sdt'], r['dv_key']), []).append(r)
    return g

g_my, g_hs = nhom(myspa_rows), nhom(hsms_rows)

khop, lech, thieu, du = [], [], [], []
for key, m_list in g_my.items():
    h_list = g_hs.get(key)
    m_con = sorted(r['con'] for r in m_list)
    if h_list is None:
        for r in m_list:
            thieu.append(r)
        continue
    h_con = sorted(r['con'] for r in h_list)
    row_info = {
        'sdt': key[0], 'ten_kh': m_list[0]['ten_kh'], 'ten_dv': m_list[0]['ten_dv'],
        'so_the_myspa': len(m_list), 'so_the_hsms': len(h_list),
        'con_myspa': '+'.join(map(str, m_con)), 'con_hsms': '+'.join(map(str, h_con)),
        'ma_the_myspa': ', '.join(r['ma_the'] for r in m_list),
        'ma_the_hsms': ', '.join(str(r.get('ma_the')) for r in h_list),
        'trang_thai_hsms': ', '.join(str(r.get('trang_thai')) for r in h_list),
    }
    if m_con == h_con:
        khop.append(row_info)
    else:
        row_info['tong_con_myspa'] = sum(m_con)
        row_info['tong_con_hsms'] = sum(h_con)
        row_info['chenh (HSMS-MySpa)'] = sum(h_con) - sum(m_con)
        lech.append(row_info)

for key, h_list in g_hs.items():
    if key not in g_my:
        for r in h_list:
            du.append(r)

# ── 4. Xuất Excel ────────────────────────────────────────────────────────────
out = os.path.join(DB_DIR, f'DOI_CHIEU_THE_SDT_DV_{datetime.now():%Y%m%d}.xlsx')
wb_out = Workbook()

def sheet(ws_, name, rows, cols):
    ws_.title = name
    ws_.append(cols)
    for r in rows:
        ws_.append([r.get(c, '') for c in cols])

COLS_NHOM = ['sdt', 'ten_kh', 'ten_dv', 'so_the_myspa', 'so_the_hsms',
             'con_myspa', 'con_hsms', 'tong_con_myspa', 'tong_con_hsms',
             'chenh (HSMS-MySpa)', 'ma_the_myspa', 'ma_the_hsms', 'trang_thai_hsms']
COLS_MY  = ['sdt', 'ten_kh', 'ten_dv', 'ma_the', 'tong', 'dung', 'con', 'cong_no']
COLS_HS  = ['sdt', 'ten_kh', 'ten_dich_vu', 'ma_the', 'so_buoi_tong', 'so_buoi_da_dung',
            'con', 'trang_thai', 'bi_dong']

sheet(wb_out.active, 'KHOP', khop, COLS_NHOM)
sheet(wb_out.create_sheet(), 'LECH_BUOI', sorted(lech, key=lambda r: -abs(r.get('chenh (HSMS-MySpa)', 0))), COLS_NHOM)
sheet(wb_out.create_sheet(), 'MYSPA_CO_HSMS_THIEU', thieu, COLS_MY)
sheet(wb_out.create_sheet(), 'HSMS_DU', du, COLS_HS)
wb_out.save(out)

n_my_ko_sdt = sum(1 for r in myspa_rows if not r['sdt'])
n_hs_ko_sdt = sum(1 for r in hsms_rows if not r['sdt'])
print(f'''
════════ KẾT QUẢ ĐỐI CHIẾU (khóa SĐT+DV+buổi, KHÔNG dùng mã thẻ) ════════
  Nhóm (SĐT,DV) khớp 100%     : {len(khop)}
  Nhóm LỆCH buổi còn lại       : {len(lech)}  ← xem tay sheet LECH_BUOI
  Thẻ MySpa có — HSMS THIẾU    : {len(thieu)} ← bổ sung trước khi cắt MySpa
  Thẻ HSMS dư (MySpa không có) : {len(du)}   (thẻ bán sau ngày export = OK)
  Bỏ qua vì không có SĐT       : MySpa {n_my_ko_sdt} · HSMS {n_hs_ko_sdt}
  📄 File: {out}
''')
