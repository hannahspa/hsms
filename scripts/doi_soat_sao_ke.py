"""Doi soat Sao Ke MB Bank vs HSMS - Doanh thu Khach CK"""
import sys
sys.stdout.reconfigure(encoding='utf-8')
import openpyxl
from supabase import create_client
from collections import defaultdict
from datetime import datetime

SUPABASE_URL = 'https://aqyemkfbjqxpegingoil.supabase.co'
SUPABASE_KEY = 'eyJhbGciOiJIUzI1NiIsInR5cCI6IkpXVCJ9.eyJpc3MiOiJzdXBhYmFzZSIsInJlZiI6ImFxeWVta2ZianF4cGVnaW5nb2lsIiwicm9sZSI6InNlcnZpY2Vfcm9sZSIsImlhdCI6MTc3NzUxNTYwMCwiZXhwIjoyMDkzMDkxNjAwfQ.L2yo4Osu6XNhPaOTEMz1Z2GI-SVtzR6AnODirhUR4zI'
supabase = create_client(SUPABASE_URL, SUPABASE_KEY)

def fmt(n):
    if n is None: return '0d'
    return f'{int(n):,}d'.replace(',', '.')

# DOC SAO KE
path = r'D:\Hannah Spa\Thu Chi\Kiem Toan\Sao Ke Quoc Nam.xlsx'
wb = openpyxl.load_workbook(path, data_only=True)
ws = wb['Giao Dịch Chi Tiết']

def parse_date(val):
    if isinstance(val, datetime):
        return val.strftime('%Y-%m-%d')
    if isinstance(val, str):
        for fmt_str in ['%d/%m/%Y', '%Y-%m-%d']:
            try: return datetime.strptime(val.strip(), fmt_str).strftime('%Y-%m-%d')
            except: pass
    return str(val) if val else '?'

transactions = []
for row in ws.iter_rows(min_row=4, max_row=ws.max_row, values_only=True):
    ngay, ma_gd, ps_no, ps_co, dien_giai, nhom, loai_ct, ghi_chu = row
    if ngay is None or (ps_no is None and ps_co is None):
        continue
    date_str = parse_date(ngay)
    no_val = int(ps_no) if ps_no else 0
    co_val = int(ps_co) if ps_co else 0
    nd = str(dien_giai).strip() if dien_giai else ''
    nhom_str = str(nhom).strip() if nhom else ''
    loai_str = str(loai_ct).strip() if loai_ct else ''
    transactions.append({
        'date': date_str, 'no': no_val, 'co': co_val,
        'nd': nd, 'nhom': nhom_str, 'loai': loai_str,
    })

# =============================================
# 1. TONG QUAN DOANH THU (THU_DV)
# =============================================
print('=' * 75)
print('DOANH THU KHACH CK — SAO KE MB BANK (12/2025 - 01/2026)')
print('=' * 75)

thu_dv = [t for t in transactions if t['loai'] == 'THU_DV' and t['co'] > 0]
print(f'Tong GD Thu DV: {len(thu_dv)}')
print(f'Tong tien: {fmt(sum(t["co"] for t in thu_dv))}')

by_month_thu = defaultdict(int)
for t in thu_dv:
    by_month_thu[t['date'][:7]] += t['co']
for m, v in sorted(by_month_thu.items()):
    n = len([t for t in thu_dv if t['date'][:7] == m])
    print(f'  {m}: {fmt(v)} ({n} GD)')

# =============================================
# 2. DOI CHIEU THEO THANG vs HSMS
# =============================================
print()
print('=' * 75)
print('DOI CHIEU THEO THANG — SAO KE vs HSMS')
print('=' * 75)

r = supabase.from_('doanh_thu').select('ngay,so_tien') \
    .eq('hinh_thuc', 'chuyen_khoan') \
    .gte('ngay', '2025-12-01').lte('ngay', '2026-01-31') \
    .order('ngay').execute()

hsms_by_month = defaultdict(int)
for d in (r.data or []):
    hsms_by_month[d['ngay'][:7]] += (d['so_tien'] or 0)

print(f'  {"Ky":<10s} {"Sao ke":>16s} {"HSMS":>16s} {"Chenh":>14s}')
for m in sorted(set(list(by_month_thu.keys()) + list(hsms_by_month.keys()))):
    sk = by_month_thu.get(m, 0)
    hs = hsms_by_month.get(m, 0)
    diff = sk - hs
    icon = 'OK' if diff == 0 else f'LECH {fmt(diff)}'
    print(f'  {m:<10s} {fmt(sk):>16s} {fmt(hs):>16s} {icon:>14s}')

total_sk = sum(by_month_thu.values())
total_hs = sum(hsms_by_month.values())
print(f'  {"-"*10} {"-"*16} {"-"*16}')
print(f'  {"TONG":<10s} {fmt(total_sk):>16s} {fmt(total_hs):>16s} {fmt(total_sk - total_hs):>14s}')

# =============================================
# 3. DOI CHIEU THEO TUNG NGAY
# =============================================
print()
print('=' * 75)
print('DOI CHIEU THEO TUNG NGAY')
print('=' * 75)

sk_by_day = defaultdict(int)
for t in thu_dv:
    sk_by_day[t['date']] += t['co']

hsms_by_day = defaultdict(int)
for d in (r.data or []):
    hsms_by_day[d['ngay']] += (d['so_tien'] or 0)

all_days = sorted(set(list(sk_by_day.keys()) + list(hsms_by_day.keys())))
print(f'  {"Ngay":<12s} {"Sao ke":>14s} {"HSMS":>14s} {"Chenh":>12s}  Status')
mismatch_days = []
for day in all_days:
    sk = sk_by_day[day]
    hs = hsms_by_day[day]
    diff = sk - hs
    status = 'OK' if diff == 0 else 'SAI'
    print(f'  {day:<12s} {fmt(sk):>14s} {fmt(hs):>14s} {fmt(diff):>12s}  {status}')
    if diff != 0:
        mismatch_days.append((day, sk, hs, diff))

print(f'\nTong ngay co GD: {len(all_days)}')
print(f'Ngay KHOP: {len(all_days) - len(mismatch_days)} / {len(all_days)}')
print(f'Ngay LECH: {len(mismatch_days)} / {len(all_days)}')

# =============================================
# 4. CHI TIET CAC NGAY SAI LECH
# =============================================
if mismatch_days:
    print()
    print('=' * 75)
    print('CHI TIET CAC NGAY SAI LECH')
    print('=' * 75)
    for day, sk, hs, diff in mismatch_days:
        print(f'\n--- {day}: Sao ke={fmt(sk)} HSMS={fmt(hs)} Chenh={fmt(diff)} ---')
        # Sao ke transactions
        sk_tx = [t for t in thu_dv if t['date'] == day]
        print(f'  SAO KE ({len(sk_tx)} GD):')
        for t in sk_tx:
            print(f'    {fmt(t["co"])}  {t["nd"][:70]}')
        # HSMS transactions
        hs_tx = supabase.from_('doanh_thu').select('so_tien,dien_giai') \
            .eq('hinh_thuc', 'chuyen_khoan').eq('ngay', day).execute()
        print(f'  HSMS ({len(hs_tx.data or [])} GD):')
        for d in (hs_tx.data or []):
            print(f'    {fmt(d["so_tien"])}  {(d.get("dien_giai") or "")[:70]}')

# =============================================
# 5. LIET KE HSMS RECORDS KHONG CO TRONG SAO KE
# =============================================
print()
print('=' * 75)
print('HSMS RECORDS KHONG CO TRONG SAO KE (co the sai hoac thua)')
print('=' * 75)
for day in all_days:
    hs = hsms_by_day.get(day, 0)
    sk = sk_by_day.get(day, 0)
    if hs > 0 and sk == 0:
        print(f'\n  {day}: HSMS co {fmt(hs)} nhung sao ke KHONG co')
        hs_tx = supabase.from_('doanh_thu').select('so_tien,dien_giai') \
            .eq('hinh_thuc', 'chuyen_khoan').eq('ngay', day).execute()
        for d in (hs_tx.data or []):
            print(f'    {fmt(d["so_tien"])}  {(d.get("dien_giai") or "")[:70]}')

print()
print('=' * 75)
print('KET LUAN')
print('=' * 75)
tong_chenh = sum(m[3] for m in mismatch_days)
print(f'Tong chenh lech: {fmt(tong_chenh)}')
print(f'Ty le khớp: {len(all_days) - len(mismatch_days)}/{len(all_days)} ngay')
