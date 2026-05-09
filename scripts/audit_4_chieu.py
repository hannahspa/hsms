"""Kiem toan 4 chieu: MySpa + Sao ke NH + HSMS"""
import sys
sys.stdout.reconfigure(encoding='utf-8')
import openpyxl
from supabase import create_client
from collections import defaultdict
from datetime import datetime

SUPABASE_URL = 'https://aqyemkfbjqxpegingoil.supabase.co'
SUPABASE_KEY = 'eyJhbGciOiJIUzI1NiIsInR5cCI6IkpXVCJ9.eyJpc3MiOiJzdXBhYmFzZSIsInJlZiI6ImFxeWVta2ZianF4cGVnaW5nb2lsIiwicm9sZSI6InNlcnZpY2Vfcm9sZSIsImlhdCI6MTc3NzUxNTYwMCwiZXhwIjoyMDkzMDkxNjAwfQ.L2yo4Osu6XNhPaOTEMz1Z2GI-SVtzR6AnODirhUR4zI'
supabase = create_client(SUPABASE_URL, SUPABASE_KEY)

def fmt(n):
    if n is None: return '0'
    return f'{int(n):,}'.replace(',', '')

# ============================================================
# 1. DOC SAO KE NGAN HANG
# ============================================================
def parse_mb(filepath):
    wb = openpyxl.load_workbook(filepath, data_only=True)
    ws = wb['Lịch sử giao dịch']
    txns = []
    for r in range(20, ws.max_row + 1):
        date_val = ws.cell(row=r, column=5).value
        debit = ws.cell(row=r, column=10).value
        credit = ws.cell(row=r, column=11).value
        desc = ws.cell(row=r, column=12).value
        if date_val is None: continue
        try:
            dt = datetime.strptime(str(date_val).strip().split(' ')[0], '%d/%m/%Y')
            date_iso = dt.strftime('%Y-%m-%d')
        except: continue
        no_val = int(float(str(debit).replace(',',''))) if debit else 0
        co_val = int(float(str(credit).replace(',',''))) if credit else 0
        nd = str(desc).strip() if desc else ''
        txns.append({'date': date_iso, 'no': no_val, 'co': co_val, 'desc': nd})
    return txns

base = r'D:\Hannah Spa\Thu Chi\Kiem Toan'
kd_bank = parse_mb(f'{base}\\KhanhDuy.xlsx')
qn_bank = parse_mb(f'{base}\\QuocNam.xlsx')

def is_khach_ck(tx):
    nd = tx['desc'].lower()
    if tx['co'] == 0: return False
    if 'tip em' in nd: return False
    if 'hoan lai' in nd or 'hoan tra' in nd or 'tra lai' in nd: return False
    if 'dt tm' in nd or 'tm dt' in nd or 'nop tien mat' in nd: return False
    if 'quy' in nd or 'quỹ' in nd: return False
    if 'lai' in nd and 'gui' in nd: return False
    return True

bank_dt_by_day = defaultdict(int)
for t in kd_bank + qn_bank:
    if is_khach_ck(t):
        bank_dt_by_day[t['date']] += t['co']

# ============================================================
# 2. DOC MYSPA DOANH THU
# ============================================================
wb_ms = openpyxl.load_workbook(f'{base}\\DoanhThuMySpa.xlsx', data_only=True)
ws_ms = wb_ms['Sheet1']

myspa_dt_by_day = defaultdict(lambda: defaultdict(int))  # day -> PTTT -> amount
myspa_total = 0
for row in ws_ms.iter_rows(min_row=2, max_row=ws_ms.max_row, values_only=True):
    ma_tt, ma_dh, ho_ten, sdt, email, ma_kh, ngay_gio, thu_ngan, ma_dh_cn, pttt, so_tien, nop_kt, ngay_kt, cn, cn_tt = row
    if so_tien is None or so_tien == 0: continue
    if isinstance(pttt, str) and 'Tổng' in pttt: continue

    date_str = str(ngay_gio).strip() if ngay_gio else ''
    try:
        dt = datetime.strptime(date_str.split(' ')[0], '%d/%m/%Y')
        date_iso = dt.strftime('%Y-%m-%d')
    except: continue

    pttt_str = str(pttt).strip() if pttt else 'Không xác định'
    amount = int(so_tien)
    myspa_dt_by_day[date_iso][pttt_str] += amount
    myspa_total += amount

# ============================================================
# 3. DOC MYSPA PHIEU CHI
# ============================================================
wb_pc = openpyxl.load_workbook(f'{base}\\PhieuChiMySpa.xlsx', data_only=True)
ws_pc = wb_pc['Sheet1']

myspa_cp_by_day = defaultdict(lambda: defaultdict(int))
myspa_cp_total = 0
for row in ws_pc.iter_rows(min_row=2, max_row=ws_pc.max_row, values_only=True):
    ma_tt, ngay_gio, phieu, nguoi, sdt, nhan, chi, pttt, loai, ly_do, nguoi_lap, cn = row
    if chi is None or chi == 0: continue
    date_str = str(ngay_gio).strip() if ngay_gio else ''
    try:
        dt = datetime.strptime(date_str.split(' ')[0], '%d/%m/%Y')
        date_iso = dt.strftime('%Y-%m-%d')
    except: continue
    pttt_str = str(pttt).strip() if pttt else 'Không xác định'
    amount = abs(int(chi))
    myspa_cp_by_day[date_iso][pttt_str] += amount
    myspa_cp_total += amount

# ============================================================
# 4. HSMS DATA
# ============================================================
r_dt = supabase.from_('doanh_thu').select('ngay,so_tien,hinh_thuc,dien_giai').order('ngay').execute()
hsms_dt_ck_by_day = defaultdict(int)
hsms_dt_tm_by_day = defaultdict(int)
hsms_dt_qt_by_day = defaultdict(int)
hsms_dt_tt_by_day = defaultdict(int)
for d in (r_dt.data or []):
    amt = d['so_tien'] or 0
    if d['hinh_thuc'] == 'chuyen_khoan':
        hsms_dt_ck_by_day[d['ngay']] += amt
    elif d['hinh_thuc'] == 'tien_mat':
        hsms_dt_tm_by_day[d['ngay']] += amt
    elif d['hinh_thuc'] == 'quet_the':
        hsms_dt_qt_by_day[d['ngay']] += amt
    elif d['hinh_thuc'] == 'the_tra_truoc':
        hsms_dt_tt_by_day[d['ngay']] += amt

r_cp = supabase.from_('chi_phi').select('ngay,so_tien,hinh_thuc_thanh_toan,dien_giai').order('ngay').execute()
hsms_cp_ck_by_day = defaultdict(int)
hsms_cp_tm_by_day = defaultdict(int)
hsms_cp_total = 0
for d in (r_cp.data or []):
    amt = d['so_tien'] or 0
    if d['hinh_thuc_thanh_toan'] == 'chuyen_khoan':
        hsms_cp_ck_by_day[d['ngay']] += amt
    elif d['hinh_thuc_thanh_toan'] == 'tien_mat':
        hsms_cp_tm_by_day[d['ngay']] += amt
    hsms_cp_total += amt

# ============================================================
# 5. SO SANH DOANH THU CHUYEN KHOAN — 3 NGUON
# ============================================================
print('=' * 90)
print('DOI CHIEU DOANH THU CHUYEN KHOAN — 3 NGUON (MySpa vs Ngan hang vs HSMS)')
print('=' * 90)

# Get MySpa DT CK by day
myspa_ck_by_day = defaultdict(int)
for day, pttt_map in myspa_dt_by_day.items():
    for pttt, amt in pttt_map.items():
        if 'chuyển khoản' in pttt.lower() or 'chuyen khoan' in pttt.lower():
            myspa_ck_by_day[day] += amt

all_days = sorted(set(list(bank_dt_by_day.keys()) + list(myspa_ck_by_day.keys()) + list(hsms_dt_ck_by_day.keys())))

print(f'\n{"Ngay":<12s} {"MySpa":>14s} {"NH Bank":>14s} {"HSMS":>14s} {"MS-NH":>12s} {"MS-HS":>12s} {"NH-HS":>12s}')
print('-' * 90)

total_ms = 0; total_nh = 0; total_hs = 0
ok_count = 0; lech_count = 0

for day in all_days:
    ms = myspa_ck_by_day.get(day, 0)
    nh = bank_dt_by_day.get(day, 0)
    hs = hsms_dt_ck_by_day.get(day, 0)
    total_ms += ms; total_nh += nh; total_hs += hs

    if ms == 0 and nh == 0 and hs == 0: continue

    ms_nh = ms - nh
    ms_hs = ms - hs
    nh_hs = nh - hs

    if ms_nh == 0 and ms_hs == 0:
        ok_count += 1
    else:
        lech_count += 1

    print(f'{day:<12s} {fmt(ms):>14s} {fmt(nh):>14s} {fmt(hs):>14s} {fmt(ms_nh):>12s} {fmt(ms_hs):>12s} {fmt(nh_hs):>12s}')

print('-' * 90)
print(f'{"TONG":<12s} {fmt(total_ms):>14s} {fmt(total_nh):>14s} {fmt(total_hs):>14s} {fmt(total_ms-total_nh):>12s} {fmt(total_ms-total_hs):>12s} {fmt(total_nh-total_hs):>12s}')

# ============================================================
# 6. SO SANH CHI PHI CHUYEN KHOAN
# ============================================================
print()
print('=' * 90)
print('DOI CHIEU CHI PHI CHUYEN KHOAN — MySpa vs HSMS')
print('=' * 90)

myspa_cp_ck_by_day = defaultdict(int)
for day, pttt_map in myspa_cp_by_day.items():
    for pttt, amt in pttt_map.items():
        if 'chuyển khoản' in pttt.lower() or 'chuyen khoan' in pttt.lower():
            myspa_cp_ck_by_day[day] += amt

cp_days = sorted(set(list(myspa_cp_ck_by_day.keys()) + list(hsms_cp_ck_by_day.keys())))
print(f'\n{"Ngay":<12s} {"MySpa CP CK":>14s} {"HSMS CP CK":>14s} {"Chenh":>12s}')
print('-' * 55)
total_ms_cp = 0; total_hs_cp = 0
for day in cp_days:
    ms = myspa_cp_ck_by_day.get(day, 0)
    hs = hsms_cp_ck_by_day.get(day, 0)
    total_ms_cp += ms; total_hs_cp += hs
    diff = ms - hs
    if ms > 0 or hs > 0:
        print(f'{day:<12s} {fmt(ms):>14s} {fmt(hs):>14s} {fmt(diff):>12s}')
print('-' * 55)
print(f'{"TONG":<12s} {fmt(total_ms_cp):>14s} {fmt(total_hs_cp):>14s} {fmt(total_ms_cp-total_hs_cp):>12s}')

# ============================================================
# 7. TONG KET
# ============================================================
print()
print('=' * 90)
print('TONG KET 4 CHIEU')
print('=' * 90)

# MySpa tong
myspa_dt_ck_total = sum(myspa_ck_by_day.values())
myspa_dt_tm_total = sum(v for day_map in myspa_dt_by_day.values() for k, v in day_map.items() if 'tiền mặt' in k.lower() or 'tien mat' in k.lower())
myspa_dt_qt_total = sum(v for day_map in myspa_dt_by_day.values() for k, v in day_map.items() if 'quẹt' in k.lower() or 'quet' in k.lower())
myspa_dt_tt_total = sum(v for day_map in myspa_dt_by_day.values() for k, v in day_map.items() if 'trả trước' in k.lower() or 'tra truoc' in k.lower())

print(f'\n  DOANH THU:')
print(f'  {"Nguon":<20s} {"CK":>16s} {"Tien mat":>16s} {"Quet the":>16s} {"The TT":>16s} {"Tong":>16s}')
print(f'  {"MySpa":<20s} {fmt(myspa_dt_ck_total):>16s} {fmt(myspa_dt_tm_total):>16s} {fmt(myspa_dt_qt_total):>16s} {fmt(myspa_dt_tt_total):>16s} {fmt(myspa_total):>16s}')

hsms_dt_ck = sum(hsms_dt_ck_by_day.values())
hsms_dt_tm = sum(hsms_dt_tm_by_day.values())
hsms_dt_qt = sum(hsms_dt_qt_by_day.values())
hsms_dt_tt = sum(hsms_dt_tt_by_day.values())
hsms_dt_all = hsms_dt_ck + hsms_dt_tm + hsms_dt_qt + hsms_dt_tt
print(f'  {"HSMS":<20s} {fmt(hsms_dt_ck):>16s} {fmt(hsms_dt_tm):>16s} {fmt(hsms_dt_qt):>16s} {fmt(hsms_dt_tt):>16s} {fmt(hsms_dt_all):>16s}')

print(f'\n  CHI PHI:')
print(f'  MySpa CP CK:  {fmt(myspa_cp_total)} (tat ca PTTT)')
print(f'  HSMS CP CK:   {fmt(hsms_cp_total)}')
print(f'  Chenh:        {fmt(myspa_cp_total - hsms_cp_total)}')

print(f'\n  NGAN HANG:')
print(f'  Tong Khach CK: {fmt(total_nh)}')
print(f'  MySpa DT CK:   {fmt(myspa_dt_ck_total)}')
print(f'  Chenh MS vs NH: {fmt(myspa_dt_ck_total - total_nh)}')
