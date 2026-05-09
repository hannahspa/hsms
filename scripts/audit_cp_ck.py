"""Kiem toan Chi Phi Chuyen Khoan: MySpa vs HSMS vs Sao ke NH"""
import sys
sys.stdout.reconfigure(encoding='utf-8')
import openpyxl
from supabase import create_client
from collections import defaultdict
from datetime import datetime

SUPABASE_URL = 'https://aqyemkfbjqxpegingoil.supabase.co'
SUPABASE_KEY = 'eyJhbGciOiJIUzI1NiIsInR5cCI6IkpXVCJ9.eyJpc3MiOiJzdXBhYmFzZSIsInJlZiI6ImFxeWVta2ZianF4cGVnaW5nb2lsIiwicm9sZSI6InNlcnZpY2Vfcm9sZSIsImlhdCI6MTc3NzUxNTYwMCwiZXhwIjoyMDkzMDkxNjAwfQ.L2yo4Osu6XNhPaOTEMz1Z2GI-SVtzR6AnODirhUR4zI'
supabase = create_client(SUPABASE_URL, SUPABASE_KEY)

def fmt(n):
    return f'{int(n):,}'.replace(',', '') if n else '0'

# ============================================================
# 1. DOC MYSPA PHIEU CHI (CHUYEN KHOAN)
# ============================================================
base = r'D:\Hannah Spa\Thu Chi\Kiem Toan'
wb = openpyxl.load_workbook(f'{base}\\PhieuChiMySpa.xlsx', data_only=True)
ws = wb['Sheet1']

myspa_cp = []
for row in ws.iter_rows(min_row=2, max_row=ws.max_row, values_only=True):
    ma_tt, ngay_gio, phieu, nguoi, sdt, nhan, chi, pttt, loai, ly_do, nguoi_lap, cn = row
    if chi is None or chi == 0: continue
    pttt_str = str(pttt).strip() if pttt else ''
    if 'chuyển khoản' not in pttt_str.lower() and 'chuyen khoan' not in pttt_str.lower(): continue

    date_str = str(ngay_gio).strip() if ngay_gio else ''
    try:
        dt = datetime.strptime(date_str.split(' ')[0], '%d/%m/%Y')
        date_iso = dt.strftime('%Y-%m-%d')
    except: continue

    myspa_cp.append({
        'date': date_iso, 'amount': abs(int(chi)),
        'loai': str(loai).strip() if loai else '',
        'ly_do': str(ly_do).strip() if ly_do else '',
        'nguoi': str(nguoi).strip() if nguoi else '',
    })

# ============================================================
# 2. DOC SAO KE NGAN HANG (CHI RA)
# ============================================================
def parse_mb(filepath):
    wb = openpyxl.load_workbook(filepath, data_only=True)
    ws = wb['Lịch sử giao dịch']
    txns = []
    for r in range(20, ws.max_row + 1):
        date_val = ws.cell(row=r, column=5).value
        debit = ws.cell(row=r, column=10).value
        credit = ws.cell(row=r, column=11).value
        desc = ws.cell(row=r, column=12).value
        if date_val is None: continue
        try:
            dt = datetime.strptime(str(date_val).strip().split(' ')[0], '%d/%m/%Y')
            date_iso = dt.strftime('%Y-%m-%d')
        except: continue
        no_val = int(float(str(debit).replace(',',''))) if debit else 0
        co_val = int(float(str(credit).replace(',',''))) if credit else 0
        nd = str(desc).strip() if desc else ''
        if no_val > 0:
            txns.append({'date': date_iso, 'amount': no_val, 'desc': nd})
    return txns

kd_bank = parse_mb(f'{base}\\KhanhDuy.xlsx')
qn_bank = parse_mb(f'{base}\\QuocNam.xlsx')
all_bank_chi = kd_bank + qn_bank

# ============================================================
# 3. DOC HSMS CHI PHI CK
# ============================================================
r = supabase.from_('chi_phi').select('*').eq('hinh_thuc_thanh_toan','chuyen_khoan').order('ngay').execute()
hsms_cp = [{'date': d['ngay'], 'amount': d['so_tien'] or 0, 'desc': d.get('dien_giai') or '', 'id': d['id']} for d in (r.data or [])]

# ============================================================
# 4. SO SANH THEO THANG
# ============================================================
print('=' * 85)
print('DOI CHIEU CHI PHI CHUYEN KHOAN — 3 NGUON')
print('=' * 85)

# Group by month
def by_month(data):
    m = defaultdict(int)
    for d in data:
        m[d['date'][:7]] += d['amount']
    return m

ms_m = by_month(myspa_cp)
hs_m = by_month(hsms_cp)
bank_m = by_month(all_bank_chi)

print(f'\n  {"Thang":<10s} {"MySpa":>16s} {"HSMS":>16s} {"NH Bank":>16s} {"MS-HS":>14s} {"HS-NH":>14s}')
months = sorted(set(list(ms_m.keys()) + list(hs_m.keys()) + list(bank_m.keys())))
for m in months:
    ms = ms_m.get(m, 0); hs = hs_m.get(m, 0); bk = bank_m.get(m, 0)
    print(f'  {m:<10s} {fmt(ms):>16s} {fmt(hs):>16s} {fmt(bk):>16s} {fmt(ms-hs):>14s} {fmt(hs-bk):>14s}')

total_ms = sum(ms_m.values()); total_hs = sum(hs_m.values()); total_bk = sum(bank_m.values())
print(f'  {\"-\"*10} {\"-\"*16} {\"-\"*16} {\"-\"*16}')
print(f'  {\"TONG\":<10s} {fmt(total_ms):>16s} {fmt(total_hs):>16s} {fmt(total_bk):>16s} {fmt(total_ms-total_hs):>14s} {fmt(total_hs-total_bk):>14s}')

# ============================================================
# 5. TIM CHI PHI CO TRONG HSMS NHUNG KHONG CO TRONG MYSPA
# ============================================================
print(f'\n{\"=\"*85}')
print('CHI PHI CK CO TRONG HSMS NHUNG KHONG CO TRONG MYSPA')
print(f'{\"=\"*85}')

# Build MySpa lookup by date+amount
ms_lookup = set()
for d in myspa_cp:
    ms_lookup.add((d['date'], d['amount']))

hsms_only = []
for d in hsms_cp:
    # Check if exists in MySpa (same date + same amount, ±100d tolerance)
    found = False
    for ms_d in myspa_cp:
        if ms_d['date'] == d['date'] and abs(ms_d['amount'] - d['amount']) <= 100:
            found = True
            break
    if not found:
        hsms_only.append(d)

print(f'\nHSMS co {len(hsms_only)} khoan chi CK KHONG co trong MySpa:')
total_hsms_only = 0
for d in sorted(hsms_only, key=lambda x: x['date']):
    total_hsms_only += d['amount']
    print(f'  {d[\"date\"]}  {fmt(d[\"amount\"]):>14s}  {d[\"desc\"][:70]}')
print(f'  Tong: {fmt(total_hsms_only)}')

# ============================================================
# 6. TIM CHI PHI CO TRONG MYSPA NHUNG KHONG CO TRONG HSMS
# ============================================================
print(f'\n{\"=\"*85}')
print('CHI PHI CK CO TRONG MYSPA NHUNG KHONG CO TRONG HSMS')
print(f'{\"=\"*85}')

ms_only = []
for ms_d in myspa_cp:
    found = False
    for hs_d in hsms_cp:
        if hs_d['date'] == ms_d['date'] and abs(hs_d['amount'] - ms_d['amount']) <= 100:
            found = True
            break
    if not found:
        ms_only.append(ms_d)

print(f'\nMySpa co {len(ms_only)} khoan chi CK KHONG co trong HSMS:')
total_ms_only = 0
for d in sorted(ms_only, key=lambda x: x['date']):
    total_ms_only += d['amount']
    print(f'  {d[\"date\"]}  {fmt(d[\"amount\"]):>14s}  {d[\"loai\"]}: {d[\"ly_do\"][:50]} ({d[\"nguoi\"]})')
print(f'  Tong: {fmt(total_ms_only)}')

# ============================================================
# 7. KET LUAN
# ============================================================
print(f'\n{\"=\"*85}')
print('KET LUAN')
print(f'{\"=\"*85}')
print(f'''
  MySpa Phiếu Chi CK:  {fmt(total_ms)} ({len(myspa_cp)} GD)
  HSMS Chi Phí CK:     {fmt(total_hs)} ({len(hsms_cp)} GD)
  Ngân hàng Chi ra:    {fmt(total_bk)} ({len(all_bank_chi)} GD)

  HSMS - MySpa = {fmt(total_hs - total_ms)} (HSMS nhiều hơn)
  -> {len(hsms_only)} khoản chi CK chỉ có trong HSMS, không có trong MySpa
  -> Đây là các khoản chi nhập thẳng vào HSMS (lương, ký quỹ, chi cá nhân...)

  MySpa - HSMS = {fmt(total_ms - total_hs)} (MySpa nhiều hơn)
  -> {len(ms_only)} khoản chi CK có trong MySpa nhưng chưa có trong HSMS
  -> Cần bổ sung vào HSMS!
''')
